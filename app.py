from dotenv import load_dotenv
load_dotenv()
from flask import Flask, render_template, request, redirect, send_file, url_for, session, jsonify
from werkzeug.utils import secure_filename
import os
import firebase_admin
from firebase_admin import credentials, db
import datetime
import io
from utils.generate_pdf import create_qc_pdf
from utils.generate_motor_qc_job_pdf import create_motor_qc_job_pdf
from utils.qr_generator import generate_qr_code
import json
import qrcode
import threading
#import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
import pdfplumber

# ✅ NEW: Cloudflare R2 (แทน Firebase Storage)
import boto3
from botocore.client import Config

app = Flask(__name__)
app.secret_key = 'your_secret_key_here'
app.config['UPLOAD_FOLDER'] = 'uploads'

# ============================================================
# 🔒 SECURITY & ALERT SYSTEM
# ============================================================
import smtplib
import time
import re
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.message import EmailMessage
from collections import defaultdict

ALERT_EMAIL        = "Chottanin@synergy-as.com"
ALERT_FROM_EMAIL   = os.environ.get("ALERT_EMAIL_ADDRESS", "")
ALERT_FROM_PASS    = os.environ.get("ALERT_EMAIL_PASSWORD", "")

# Rate limiting — กัน brute force login
_login_attempts    = defaultdict(list)   # ip -> [timestamps]
MAX_ATTEMPTS       = 5                   # ครั้งสูงสุดต่อ 10 นาที
ATTEMPT_WINDOW     = 600                 # วินาที
BLOCKED_IPS        = set()

# Suspicious path patterns
SUSPICIOUS_PATTERNS = [
    r"\.\./", r"\.\.\\",
    r"(etc/passwd|etc/shadow|proc/self)",
    r"(wp-admin|wp-login|phpMyAdmin|\.php)",
    r"(eval\(|base64_decode|exec\(|system\()",
    r"(<script|javascript:|on\w+=)",
    r"(UNION.+SELECT|SELECT.+FROM|DROP.+TABLE)",
    r"(\x00|%00)",
]
_suspicious_re = [re.compile(p, re.IGNORECASE) for p in SUSPICIOUS_PATTERNS]

_pending_alerts    = []
_alert_lock        = threading.Lock()
_last_alert_sent   = 0
ALERT_COOLDOWN     = 300  # ส่งอีเมลซ้ำได้ทุก 5 นาที


def _send_alert_email(subject: str, body: str):
    if not ALERT_FROM_EMAIL or not ALERT_FROM_PASS:
        print(f"[ALERT] {subject}: {body}", flush=True)
        return
    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = f"[SAS QC Alert] {subject}"
        msg["From"]    = ALERT_FROM_EMAIL
        msg["To"]      = ALERT_EMAIL
        html = f"""
        <html><body style="font-family:sans-serif;padding:20px;">
        <h2 style="color:#cc0000;">⚠️ SAS QC System Alert</h2>
        <table border="0" cellpadding="8" style="border-collapse:collapse;width:100%;">
        <tr><td style="background:#f5f5f5;font-weight:bold;">เรื่อง</td>
            <td>{subject}</td></tr>
        <tr><td style="background:#f5f5f5;font-weight:bold;">รายละเอียด</td>
            <td><pre style="margin:0;">{body}</pre></td></tr>
        <tr><td style="background:#f5f5f5;font-weight:bold;">เวลา (UTC+7)</td>
            <td>{(datetime.datetime.utcnow()+datetime.timedelta(hours=7)).strftime('%Y-%m-%d %H:%M:%S')}</td></tr>
        </table>
        <p style="color:#888;font-size:12px;">SAS QC Gearmotor — Automated Security Alert</p>
        </body></html>"""
        msg.attach(MIMEText(html, "html"))
        with smtplib.SMTP_SSL("smtp.gmail.com", 465, timeout=10) as smtp:
            smtp.login(ALERT_FROM_EMAIL, ALERT_FROM_PASS)
            smtp.sendmail(ALERT_FROM_EMAIL, ALERT_EMAIL, msg.as_string())
        print(f"[ALERT SENT] {subject}", flush=True)
    except Exception as e:
        print(f"[ALERT ERROR] {e}", flush=True)


def _flush_alerts():
    global _last_alert_sent
    with _alert_lock:
        if not _pending_alerts:
            return
        now = time.time()
        if now - _last_alert_sent < ALERT_COOLDOWN:
            return
        events = list(_pending_alerts)
        _pending_alerts.clear()
        _last_alert_sent = now
    body = "\n\n---\n\n".join(events)
    threading.Thread(
        target=_send_alert_email,
        args=(f"{len(events)} Security Event(s) Detected", body),
        daemon=True
    ).start()


def _queue_alert(event: str):
    with _alert_lock:
        _pending_alerts.append(event)
    threading.Thread(target=_flush_alerts, daemon=True).start()


def _get_client_ip():
    return (request.headers.get("X-Forwarded-For","") or "").split(",")[0].strip() \
           or request.remote_addr or "unknown"


def _is_suspicious_request() -> bool:
    check_str = request.full_path + " " + str(request.data)[:500]
    return any(rx.search(check_str) for rx in _suspicious_re)


@app.before_request
def security_middleware():
    ip = _get_client_ip()
    if ip in BLOCKED_IPS:
        _queue_alert(
            f"BLOCKED IP tried again\nIP: {ip}\nPath: {request.path}"
        )
        return "Forbidden", 403
    if not request.path.startswith("/static/"):
        if _is_suspicious_request():
            BLOCKED_IPS.add(ip)
            _queue_alert(
                f"SUSPICIOUS REQUEST — IP auto-blocked\n"
                f"IP: {ip}\nPath: {request.full_path}\n"
                f"UA: {request.headers.get('User-Agent','')[:200]}"
            )
            return "Forbidden", 403


@app.after_request
def security_headers(response):
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"]         = "SAMEORIGIN"
    response.headers["X-XSS-Protection"]        = "1; mode=block"
    response.headers["Referrer-Policy"]         = "strict-origin-when-cross-origin"
    return response


def check_login_rate_limit(ip: str) -> bool:
    now = time.time()
    attempts = [t for t in _login_attempts[ip] if t > now - ATTEMPT_WINDOW]
    _login_attempts[ip] = attempts
    if len(attempts) >= MAX_ATTEMPTS:
        return False
    _login_attempts[ip].append(now)
    return True


def record_failed_login(ip: str, employee_id: str):
    now = time.time()
    attempts = [t for t in _login_attempts[ip] if t > now - ATTEMPT_WINDOW]
    if len(attempts) >= MAX_ATTEMPTS - 1:
        _queue_alert(
            f"BRUTE FORCE LOGIN\nIP: {ip}\n"
            f"ID ที่ลอง: {employee_id}\n"
            f"พยายามใน 10 นาที: {len(attempts)+1} ครั้ง"
        )


def _send_startup_alert():
    time.sleep(8)
    _send_alert_email(
        "Server Started",
        f"SAS QC server เริ่มทำงานแล้ว\n"
        f"เวลา (UTC+7): {(datetime.datetime.utcnow()+datetime.timedelta(hours=7)).strftime('%Y-%m-%d %H:%M:%S')}\n"
        f"URL: https://sas-qc-gearmotor.onrender.com"
    )

threading.Thread(target=_send_startup_alert, daemon=True).start()

# ==== Load Firebase Credential from Environment ====
# (ยังคงใช้ Firebase Realtime Database เหมือนเดิมทุกอย่าง)
firebase_cred_str = os.environ.get("FIREBASE_CREDENTIAL_JSON")
if not firebase_cred_str:
    raise ValueError("❌ FIREBASE_CREDENTIAL_JSON env variable is missing!")
firebase_json = json.loads(firebase_cred_str)
cred = credentials.Certificate(firebase_json)

firebase_admin.initialize_app(cred, {
    'databaseURL': 'https://sas-qc-gearmotor-app-default-rtdb.asia-southeast1.firebasedatabase.app/',
})

ref = db.reference("/qc_data")

# ==== Cloudflare R2 Setup ====
# ตั้งค่า Environment Variables บน Render.com:
#   R2_ACCOUNT_ID     = Cloudflare Account ID (ดูได้จาก R2 dashboard)
#   R2_ACCESS_KEY_ID  = R2 API Token Access Key ID
#   R2_SECRET_KEY     = R2 API Token Secret Access Key
#   R2_BUCKET_NAME    = ชื่อ bucket ที่สร้างไว้
#   R2_PUBLIC_URL     = https://<your-bucket>.<your-subdomain>.r2.dev  (หรือ custom domain)
#
# วิธีสร้าง R2 API Token:
#   1. เข้า Cloudflare Dashboard → R2 → Manage R2 API Tokens
#   2. Create API Token → ให้สิทธิ์ Object Read & Write
#   3. คัดลอก Access Key ID และ Secret Access Key ไปใส่ใน Render Environment

R2_ACCOUNT_ID    = os.environ.get("R2_ACCOUNT_ID", "")
R2_ACCESS_KEY_ID = os.environ.get("R2_ACCESS_KEY_ID", "")
R2_SECRET_KEY    = os.environ.get("R2_SECRET_KEY", "")
R2_BUCKET_NAME   = os.environ.get("R2_BUCKET_NAME", "sas-qc-gearmotor")
R2_PUBLIC_URL    = os.environ.get("R2_PUBLIC_URL", "").rstrip("/")

r2 = boto3.client(
    "s3",
    endpoint_url=f"https://{R2_ACCOUNT_ID}.r2.cloudflarestorage.com",
    aws_access_key_id=R2_ACCESS_KEY_ID,
    aws_secret_access_key=R2_SECRET_KEY,
    config=Config(signature_version="s3v4"),
    region_name="auto",
)

def r2_upload_fileobj(fileobj, key, content_type):
    """
    อัปโหลด file-like object ขึ้น R2
    คืน public URL ของไฟล์
    """
    r2.upload_fileobj(
        fileobj,
        R2_BUCKET_NAME,
        key,
        ExtraArgs={"ContentType": content_type},
    )
    return f"{R2_PUBLIC_URL}/{key}"

def r2_upload_bytes(data_bytes, key, content_type):
    """
    อัปโหลด bytes ขึ้น R2
    คืน public URL ของไฟล์
    """
    r2.put_object(
        Bucket=R2_BUCKET_NAME,
        Key=key,
        Body=data_bytes,
        ContentType=content_type,
    )
    return f"{R2_PUBLIC_URL}/{key}"

def r2_download_bytes(key):
    """
    ดาวน์โหลดไฟล์จาก R2 คืนเป็น bytes
    """
    resp = r2.get_object(Bucket=R2_BUCKET_NAME, Key=key)
    return resp["Body"].read()

def r2_get_mtime(key):
    """
    ดึง LastModified ของ object ใน R2 (ใช้เป็น cache key เหมือนเดิม)
    """
    resp = r2.head_object(Bucket=R2_BUCKET_NAME, Key=key)
    return str(resp.get("LastModified", ""))


# ============================================================
# 🎯 Gear Motor Matcher Integration
# ============================================================
# Import และเปิดใช้งาน Matcher blueprint
# ใช้ R2 client เดิม + security middleware เดิมของระบบ
# ต้องตั้ง env var: ANTHROPIC_API_KEY
from utils.gear_matcher import init_matcher, matcher_bp

init_matcher(
    r2_client=r2,
    bucket=R2_BUCKET_NAME,
    public_url=R2_PUBLIC_URL,
    anthropic_key=os.environ.get("ANTHROPIC_API_KEY", ""),
)
app.register_blueprint(matcher_bp, url_prefix='/matcher')

# =========================
# Stock on shelf (FAST + SIMPLE)
# =========================
STOCK_BLOB_NAME = "stock/Stock motor.xlsx"
STOCK_XLS_PATH = os.path.join(os.path.dirname(__file__), "Stock motor.xlsx")
STOCK_UPLOAD_PASS = "Adminsas2026"

# ✅ Supplier (ZD) stock file
SUPPLIER_BLOB_NAME = "stock/Stock ZD.xlsx"
SUPPLIER_UPLOAD_PASS = "Chottaninsas2029"

CHECK_BLOB_NAME = "stock/Check status.xlsx"
CHECK_UPLOAD_PASS = "Adminsas2026"
_check_cache = {"mtime": None, "rows": []}
_check_lock = threading.Lock()

_stock_cache = {
    "mtime": None,
    "rows": []
}
_stock_lock = threading.Lock()

# ✅ cache for Supplier ZD
_supplier_cache = {"mtime": None, "rows": []}
_supplier_lock = threading.Lock()


def _load_stock_rows_cached():
    """
    Read latest stock xlsx from R2 and cache by object mtime.
    Range:
      Code: A (col 1)
      Description: E (col 5)
      Total: AF (col 32)
    """
    mtime = r2_get_mtime(STOCK_BLOB_NAME)

    with _stock_lock:
        if _stock_cache["mtime"] == mtime and _stock_cache["rows"] is not None:
            return _stock_cache["rows"]

        data = r2_download_bytes(STOCK_BLOB_NAME)
        wb = load_workbook(BytesIO(data), data_only=True, read_only=True)
        ws = next((s for s in wb.worksheets if (s.max_column or 0) >= 32 and (s.max_row or 0) >= 3000), wb.worksheets[0])

        rows = []

        for row in ws.iter_rows(min_row=2, max_row=3000, min_col=1, max_col=32, values_only=True):
            code  = row[0] if len(row) > 0 else None   # A
            desc  = row[4] if len(row) > 4 else None   # E
            total = row[31] if len(row) > 31 else None  # AF

            code_s = "" if code is None else str(code).strip()
            if code_s:
                rows.append({
                    "code": code_s,
                    "description": "" if desc is None else str(desc).strip(),
                    "total": "" if total is None else str(total).strip()
                })

        _stock_cache["mtime"] = mtime
        _stock_cache["rows"] = rows
        return rows


# ✅ Supplier ZD rows (Sheet1, A2-A400, B2-B400)
def _load_supplier_rows_cached():
    mtime = r2_get_mtime(SUPPLIER_BLOB_NAME)

    with _supplier_lock:
        if _supplier_cache["mtime"] == mtime and _supplier_cache["rows"] is not None:
            return _supplier_cache["rows"]

        data = r2_download_bytes(SUPPLIER_BLOB_NAME)
        wb = load_workbook(BytesIO(data), data_only=True, read_only=True)

        ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.worksheets[0]

        rows = []
        for row in ws.iter_rows(min_row=2, max_row=400, min_col=1, max_col=2, values_only=True):
            code = row[0] if len(row) > 0 else None  # A
            total = row[1] if len(row) > 1 else None  # B

            code_s = "" if code is None else str(code).strip()
            if code_s:
                rows.append({
                    "code": code_s,
                    "description": "Supplier",
                    "total": "" if total is None else str(total).strip()
                })

        _supplier_cache["mtime"] = mtime
        _supplier_cache["rows"] = rows
        return rows

    
def _load_check_rows_cached():
    mtime = r2_get_mtime(CHECK_BLOB_NAME)

    with _check_lock:
        if _check_cache["mtime"] == mtime and _check_cache["rows"] is not None:
            return _check_cache["rows"]

        data = r2_download_bytes(CHECK_BLOB_NAME)
        wb = load_workbook(BytesIO(data), data_only=True, read_only=True)
        year_sheets = [s for s in wb.sheetnames if str(s).isdigit()]
        ws = wb[str(max(map(int, year_sheets)))] if year_sheets else wb.worksheets[0]

        rows = []

        for row in ws.iter_rows(min_row=25, max_row=600, min_col=1, max_col=18, values_only=True):
            no_item      = row[0]   # A
            po_no        = row[3]   # D
            po_open_date = row[4]   # E
            stock_order  = row[5]   # G
            amount       = row[6]   # H
            transport    = row[12]  # N
            factory_eta  = row[15]  # Q
            delivery_eta = row[16]  # R

            if (no_item is None and po_no is None and po_open_date is None
                and stock_order is None and amount is None and transport is None
                and factory_eta is None and delivery_eta is None):
                continue

            rows.append({
                "no_item": "" if no_item is None else str(no_item).strip(),
                "po_open_date": "" if po_open_date is None else str(po_open_date).strip(),
                "po": "" if po_no is None else str(po_no).strip(),
                "stock_or_order": "" if stock_order is None else str(stock_order).strip(),
                "amount": "" if amount is None else str(amount).strip(),
                "transport": "" if transport is None else str(transport).strip(),
                "factory_eta": "" if factory_eta is None else str(factory_eta).strip(),
                "delivery_eta": "" if delivery_eta is None else str(delivery_eta).strip(),
            })

        _check_cache["mtime"] = mtime
        _check_cache["rows"] = rows
        return rows

    
@app.route("/stock-upload", methods=["GET", "POST"])
def stock_upload_public():
    if request.method == "GET":
        key = (request.args.get("key") or "").strip()
        if key != STOCK_UPLOAD_PASS:
            return "Unauthorized", 403

        session["stock_upload_ok"] = True
        return render_template("stock_upload_public.html")

    # ===== POST =====
    if not session.get("stock_upload_ok"):
        return "Unauthorized", 403

    f = request.files.get("file")
    if not f or f.filename.strip() == "":
        return "No file uploaded", 400

    if not f.filename.lower().endswith(".xlsx"):
        return "Invalid file type (ต้องเป็น .xlsx เท่านั้น)", 400

    # ✅ อัปโหลดขึ้น R2 แทน Firebase Storage
    r2_upload_fileobj(
        f,
        STOCK_BLOB_NAME,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # ล้าง cache เพื่อให้หน้า stock ดึงไฟล์ใหม่ทันที
    with _stock_lock:
        _stock_cache["mtime"] = None
        _stock_cache["rows"] = []

    session.pop("stock_upload_ok", None)

    return redirect("/stock")


@app.route("/check-status-upload", methods=["GET", "POST"])
def check_status_upload_public():
    if request.method == "GET":
        key = (request.args.get("key") or "").strip()
        if key != CHECK_UPLOAD_PASS:
            return "Unauthorized", 403

        session["check_upload_ok"] = True
        return render_template("check_status_upload_public.html")

    # ===== POST =====
    if not session.get("check_upload_ok"):
        return "Unauthorized", 403

    f = request.files.get("file")
    if not f or f.filename.strip() == "":
        return "No file uploaded", 400

    if not f.filename.lower().endswith(".xlsx"):
        return "Invalid file type (ต้องเป็น .xlsx เท่านั้น)", 400

    # ✅ อัปโหลดขึ้น R2 แทน Firebase Storage
    r2_upload_fileobj(
        f,
        CHECK_BLOB_NAME,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # ล้าง cache
    with _check_lock:
        _check_cache["mtime"] = None
        _check_cache["rows"] = []

    session.pop("check_upload_ok", None)

    return redirect("/check-status")

                
# ✅ Inspector mapping (ID -> ชื่อ)
INSPECTOR_MAP = {
    "QC001": "คุณประสิทธิ์",
    "QC002": "คุณเกียรติศักดิ์",
    "QC003": "คุณมัด",
    "QC999": "คุณโชติธนินท์",
}

# ✅ Product type list ใช้ร่วมกันระหว่าง form.html และหน้าออกเอกสาร QC-Motor
PRODUCT_TYPES = [
    "AC/DC Gear Motor",
    "BLDC Gear Motor",
    "Servo Motor and Servo Drive",
    "Planetary GearReducer",
    "RFKS Series",
    "H/B Gear Box",
    "Hypoid GearMotor",
    "Cycloidal GearMotor",
    "SRV WormGear",
    "SMR Gearbox",
    "Small AC Gearmotor",
    "Belt Conveyor",
    "Automation",
    "Others",
]

# ✅ Admin Motor QC Job Database
motor_qc_jobs_ref = db.reference("/motor_qc_jobs")


def _safe_firebase_key(value: str) -> str:
    """Firebase key ห้ามมี . # $ [ ] / จึงต้องแปลงให้ปลอดภัย"""
    return re.sub(r"[.#$\[\]/]", "_", str(value or "").strip())[:120]


def _safe_next_url(next_url: str) -> str:
    next_url = str(next_url or "").strip()
    if next_url.startswith("/") and not next_url.startswith("//"):
        return next_url
    return ""


def _next_motor_qc_no():
    prefix = "MO" + datetime.datetime.now().strftime("%Y%m%d")
    try:
        rows = motor_qc_jobs_ref.order_by_key().start_at(prefix).end_at(prefix + "\uf8ff").get() or {}
        max_no = 0
        if isinstance(rows, dict):
            for key in rows.keys():
                m = re.match(rf"^{prefix}(\d{{3}})$", str(key))
                if m:
                    max_no = max(max_no, int(m.group(1)))
        return prefix + str(max_no + 1).zfill(3)
    except Exception:
        return prefix + "001"


def _product_type_summary(items):
    types = []
    for item in items or []:
        t = str((item or {}).get('product_type') or '').strip()
        if t and t not in types:
            types.append(t)
    if not types:
        return ''
    if len(types) == 1:
        return types[0]
    return 'หลายประเภทสินค้า'


def _collect_qc_report_images(report_data):
    """รวมรูปสำหรับสร้าง PDF รองรับทั้ง legacy images และ qc_items.images"""
    report_data = report_data or {}
    image_urls = []
    image_labels = []

    item_order = [
        ('rfks_nameplate_motor_img', 'Name plate : Motor'),
        ('rfks_nameplate_gear_img', 'Name plate : Gear'),
        ('motor_current_img', 'ภาพค่ากระแส'),
        ('gear_sound_img', 'ภาพเสียงเกียร์ / Run Test'),
        ('assembly_img', 'ภาพประกอบหน้างาน'),
        ('controller_img', 'ภาพ Controller'),
        ('servo_motor_img', 'ภาพ Servo Motor'),
        ('servo_drive_img', 'ภาพ Servo Drive'),
        ('cable_wire_img', 'ภาพ Cable Wire'),
    ]

    qc_items = report_data.get('qc_items') or []
    if isinstance(qc_items, list) and qc_items:
        for item in qc_items:
            images = item.get('images') or {}
            no = item.get('no', '')
            model = str(item.get('model') or '').strip()
            model_part = f" - {model[:42]}" if model else ''
            for key, label in item_order:
                url = images.get(key)
                if url:
                    image_urls.append(url)
                    image_labels.append(f"Item {no}{model_part} | {label}")
        return image_urls, image_labels

    images_dict = report_data.get('images', {}) or {}
    legacy_order = [
        ('rfks_nameplate_motor_img', 'Name plate : Motor'),
        ('rfks_nameplate_gear_img', 'Name plate : Gear'),
        ('motor_current_img', 'ภาพค่ากระแส'),
        ('gear_sound_img', 'ภาพเสียงเกียร์'),
        ('assembly_img', 'ภาพประกอบหน้างาน'),
        ('controller_img', 'ภาพ Controller'),
        ('servo_motor_img', 'ภาพ Servo Motor'),
        ('servo_drive_img', 'ภาพ Servo Drive'),
        ('cable_wire_img', 'ภาพ Cable Wire'),
    ]
    for key, label in legacy_order:
        url = images_dict.get(key)
        if url:
            image_urls.append(url)
            image_labels.append(label)
    return image_urls, image_labels


# ============================================================
# ✅ Motor QC Approval Workflow Helpers
# ============================================================
def _make_qr_stream(data: str, box_size: int = 6):
    """สร้าง QR PNG stream จาก URL/ข้อความ"""
    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=box_size,
        border=2,
    )
    qr.add_data(str(data or ''))
    qr.make(fit=True)
    img = qr.make_image(fill_color='black', back_color='white')
    stream = io.BytesIO()
    img.save(stream, format='PNG')
    stream.seek(0)
    return stream


def _motor_qc_approval_urls(job_key: str):
    safe_key = _safe_firebase_key(job_key)
    return {
        'warehouse': url_for('motor_qc_department_approve', role='warehouse', job_key=safe_key, _external=True),
        'qc': url_for('motor_qc_department_approve', role='qc', job_key=safe_key, _external=True),
    }


def _motor_qc_logo_path():
    logo_path = os.path.join(app.root_path, 'static', 'logo_sas.png')
    return logo_path if os.path.exists(logo_path) else None


def _pdf_reader_writer():
    """โหลด PDF library เฉพาะตอนใช้งาน เพื่อให้ระบบเดิมยัง start ได้"""
    try:
        from pypdf import PdfReader, PdfWriter
        return PdfReader, PdfWriter
    except Exception as e:
        raise RuntimeError(
            "ต้องติดตั้ง pypdf เพื่อรวมไฟล์ใบจองเข้ากับ QC-GEARMOTOR PRE-CHECK DOCUMENT "
            "ให้เพิ่ม pypdf ใน requirements.txt แล้ว Deploy ใหม่"
        ) from e


def _pdf_page_count(pdf_bytes: bytes) -> int:
    PdfReader, _PdfWriter = _pdf_reader_writer()
    reader = PdfReader(io.BytesIO(pdf_bytes or b''))
    if getattr(reader, 'is_encrypted', False):
        try:
            reader.decrypt('')
        except Exception:
            raise RuntimeError('ไฟล์ PDF ถูกเข้ารหัส ไม่สามารถแนบ/รวมไฟล์ได้')
    return len(reader.pages)


def _merge_pdf_bytes(pdf_parts) -> bytes:
    """รวม PDF หลายชุดให้อยู่ในไฟล์เดียว: PRE-CHECK + ใบจอง"""
    PdfReader, PdfWriter = _pdf_reader_writer()
    writer = PdfWriter()
    for idx, part in enumerate(pdf_parts or [], start=1):
        if not part:
            continue
        reader = PdfReader(io.BytesIO(part))
        if getattr(reader, 'is_encrypted', False):
            try:
                reader.decrypt('')
            except Exception:
                raise RuntimeError(f'ไฟล์ PDF ลำดับที่ {idx} ถูกเข้ารหัส ไม่สามารถรวมไฟล์ได้')
        for page in reader.pages:
            writer.add_page(page)
    out = io.BytesIO()
    writer.write(out)
    return out.getvalue()


def _append_booking_pdf_if_any(precheck_pdf_bytes: bytes, job: dict) -> bytes:
    """ถ้ามีใบจอง ให้รวมท้าย PRE-CHECK PDF ทุกครั้งก่อน download/email/upload"""
    job = job or {}
    booking_key = str(job.get('booking_pdf_key') or '').strip()
    if not booking_key:
        return precheck_pdf_bytes
    try:
        booking_bytes = r2_download_bytes(booking_key)
        merged = _merge_pdf_bytes([precheck_pdf_bytes, booking_bytes])
        print(f"[MOTOR QC PDF] merged booking PDF key={booking_key} pages={job.get('booking_pdf_page_count','?')}", flush=True)
        return merged
    except Exception as e:
        raise RuntimeError(f"รวมไฟล์ใบจองเข้ากับ QC-GEARMOTOR PRE-CHECK DOCUMENT ไม่สำเร็จ: {e}")


def _build_motor_qc_precheck_pdf_bytes(job: dict) -> bytes:
    """สร้าง PDF QC-GEARMOTOR PRE-CHECK ล่าสุด พร้อม QR/ลายเซ็น และแนบใบจองท้ายไฟล์ถ้ามี"""
    job = job or {}
    job_key = job.get('job_key') or job.get('qr_no') or ''
    approval_urls = job.get('approval_urls') or _motor_qc_approval_urls(job_key)
    job['approval_urls'] = approval_urls

    form_url = job.get('form_url') or url_for('form', motor_qc_job=_safe_firebase_key(job_key), _external=True)
    job['form_url'] = form_url

    pdf_stream = create_motor_qc_job_pdf(
        job,
        qr_image_stream=_make_qr_stream(form_url, box_size=8),
        barcode_value=form_url,
        logo_path=_motor_qc_logo_path(),
        qc_qr_image_stream=_make_qr_stream(approval_urls.get('qc'), box_size=5),
        warehouse_qr_image_stream=_make_qr_stream(approval_urls.get('warehouse'), box_size=5),
    )
    pdf_stream.seek(0)
    precheck_bytes = pdf_stream.read()
    return _append_booking_pdf_if_any(precheck_bytes, job)


def _env_first(*names):
    """คืนค่า env ตัวแรกที่มีค่า พร้อมชื่อ env ที่ใช้จริง"""
    for name in names:
        val = os.environ.get(name)
        if val is not None and str(val).strip():
            return str(val).strip(), name
    return '', ''


def _workflow_mail_config():
    """อ่าน SMTP config แบบรองรับชื่อ env หลายแบบ เพื่อไม่ให้ Production ข้ามการส่งเมลเงียบ ๆ"""
    smtp_user, user_env = _env_first(
        'SMTP_EMAIL_ADDRESS', 'SMTP_USERNAME', 'SMTP_USER',
        'ALERT_EMAIL_ADDRESS', 'EMAIL_ADDRESS', 'MAIL_USERNAME',
        'GMAIL_USER', 'GMAIL_EMAIL'
    )
    smtp_pass, pass_env = _env_first(
        'SMTP_EMAIL_PASSWORD', 'SMTP_PASSWORD', 'SMTP_PASS',
        'ALERT_EMAIL_PASSWORD', 'EMAIL_PASSWORD', 'MAIL_PASSWORD',
        'GMAIL_APP_PASSWORD', 'GMAIL_PASSWORD'
    )
    smtp_host, host_env = _env_first('SMTP_HOST', 'MAIL_SERVER')
    smtp_port_raw, port_env = _env_first('SMTP_PORT', 'MAIL_PORT')

    smtp_host = smtp_host or 'smtp.gmail.com'
    try:
        smtp_port = int(str(smtp_port_raw or '465').strip())
    except Exception:
        smtp_port = 465

    # Gmail App Password มักคัดลอกมาแบบมีช่องว่าง เช่น abcd efgh ijkl mnop
    # ต้องส่ง login แบบไม่มีช่องว่าง
    if 'gmail' in smtp_host.lower():
        smtp_pass = re.sub(r'\s+', '', smtp_pass or '')

    return {
        'user': smtp_user,
        'password': smtp_pass,
        'host': smtp_host,
        'port': smtp_port,
        'user_env': user_env,
        'pass_env': pass_env,
        'host_env': host_env or 'default',
        'port_env': port_env or 'default',
    }


def _record_mail_event(job_key, stage, result):
    """บันทึกสถานะการส่งเมลไว้ใน Firebase เพื่อเปิดดูย้อนหลังได้ว่าเมล fail เพราะอะไร"""
    if not job_key:
        return
    try:
        safe_key = _safe_firebase_key(job_key)
        now_th = datetime.datetime.utcnow() + datetime.timedelta(hours=7)
        payload = {
            'stage': stage or '',
            'ok': bool(result.get('ok')),
            'subject': result.get('subject', ''),
            'to': result.get('to', []),
            'cc': result.get('cc', []),
            'error': result.get('error', ''),
            'smtp_host': result.get('smtp_host', ''),
            'smtp_port': result.get('smtp_port', ''),
            'smtp_user_env': result.get('smtp_user_env', ''),
            'smtp_pass_env': result.get('smtp_pass_env', ''),
            'created_at': now_th.strftime('%Y-%m-%d %H:%M:%S'),
        }
        motor_qc_jobs_ref.child(safe_key).child('mail_events').push(payload)
        motor_qc_jobs_ref.child(safe_key).child('last_mail_status').set(payload)
    except Exception as e:
        print(f"[MAIL LOG ERROR] {e}", flush=True)


def _send_workflow_email(to_list, cc_list, subject, body, attachment_bytes=None, attachment_name=None, stage='', job_key=''):
    """ส่งอีเมล Workflow พร้อมแนบ PDF โดยใช้ SMTP ของระบบ ไม่ผูกกับเครื่องผู้ใช้

    คืนค่า dict เสมอ เพื่อ debug Production ได้ ไม่ข้ามเงียบเหมือน V5
    """
    cfg = _workflow_mail_config()

    to_list = [str(x).strip() for x in (to_list or []) if str(x).strip()]
    cc_list = [str(x).strip() for x in (cc_list or []) if str(x).strip()]
    all_recipients = to_list + cc_list

    result = {
        'ok': False,
        'stage': stage or '',
        'subject': subject or '',
        'to': to_list,
        'cc': cc_list,
        'smtp_host': cfg['host'],
        'smtp_port': cfg['port'],
        'smtp_user_env': cfg['user_env'],
        'smtp_pass_env': cfg['pass_env'],
        'error': '',
    }

    if not cfg['user'] or not cfg['password']:
        result['error'] = (
            'SMTP env missing: ให้ตั้งค่า SMTP_EMAIL_ADDRESS และ SMTP_EMAIL_PASSWORD '
            'บน Render Environment หรือใช้ ALERT_EMAIL_ADDRESS / ALERT_EMAIL_PASSWORD เป็น fallback ได้'
        )
        print(f"[MAIL FAILED] {result['error']} | subject={subject}", flush=True)
        _record_mail_event(job_key, stage, result)
        return result

    if not all_recipients:
        result['error'] = 'No recipient email address'
        print(f"[MAIL FAILED] {result['error']} | subject={subject}", flush=True)
        _record_mail_event(job_key, stage, result)
        return result

    try:
        msg = EmailMessage()
        msg['Subject'] = subject
        msg['From'] = f"SAS QC Motor <{cfg['user']}>"
        msg['To'] = ', '.join(to_list)
        if cc_list:
            msg['Cc'] = ', '.join(cc_list)
        msg.set_content(body or '')

        if attachment_bytes and attachment_name:
            msg.add_attachment(
                attachment_bytes,
                maintype='application',
                subtype='pdf',
                filename=attachment_name,
            )

        if int(cfg['port']) == 465:
            with smtplib.SMTP_SSL(cfg['host'], int(cfg['port']), timeout=30) as smtp:
                smtp.login(cfg['user'], cfg['password'])
                smtp.send_message(msg, from_addr=cfg['user'], to_addrs=all_recipients)
        else:
            with smtplib.SMTP(cfg['host'], int(cfg['port']), timeout=30) as smtp:
                smtp.ehlo()
                smtp.starttls()
                smtp.ehlo()
                smtp.login(cfg['user'], cfg['password'])
                smtp.send_message(msg, from_addr=cfg['user'], to_addrs=all_recipients)

        result['ok'] = True
        print(f"[MAIL SENT] stage={stage} | to={to_list} | cc={cc_list} | subject={subject}", flush=True)
        _record_mail_event(job_key, stage, result)
        return result

    except smtplib.SMTPAuthenticationError as e:
        result['error'] = f"SMTP Authentication failed: ตรวจ Gmail App Password / 2-Step Verification / Email ผู้ส่ง ({e})"
    except smtplib.SMTPRecipientsRefused as e:
        result['error'] = f"SMTP recipients refused: {e}"
    except smtplib.SMTPException as e:
        result['error'] = f"SMTP error: {e}"
    except Exception as e:
        result['error'] = f"Unexpected mail error: {e}"

    print(f"[MAIL ERROR] stage={stage} | {result['error']} | subject={subject}", flush=True)
    _record_mail_event(job_key, stage, result)
    return result


@app.route('/admin-motor-qc/mail-test')
def admin_motor_qc_mail_test():
    """ทดสอบ SMTP บน Production โดยต้องตั้ง MAIL_TEST_KEY ก่อน เช่น /admin-motor-qc/mail-test?key=xxxx&to=email"""
    required_key = (os.environ.get('MAIL_TEST_KEY') or '').strip()
    supplied_key = (request.args.get('key') or '').strip()
    if not required_key or supplied_key != required_key:
        return jsonify({
            'ok': False,
            'error': 'Unauthorized หรือยังไม่ได้ตั้ง MAIL_TEST_KEY บน Render Environment'
        }), 403

    cfg = _workflow_mail_config()
    test_to = (request.args.get('to') or cfg.get('user') or 'Chottanin@synergy-as.com').strip()
    now_th = datetime.datetime.utcnow() + datetime.timedelta(hours=7)
    res = _send_workflow_email(
        to_list=[test_to],
        cc_list=[],
        subject=f"SAS QC Motor Mail Test {now_th.strftime('%Y-%m-%d %H:%M:%S')}",
        body="ทดสอบการส่งอีเมลจาก Production Render ของระบบ SAS QC Motor",
        stage='mail_test',
        job_key='',
    )
    status = 200 if res.get('ok') else 500
    return jsonify(res), status


def _motor_qc_checkpoints():
    """หัวข้อ QC ที่ QC Inspector ต้องติ๊กก่อน Approve และจะถูกพิมพ์ X ลง PDF ล่าสุด"""
    return [
        ('1', 'ตรวจจำนวนสินค้า / Model / Nameplate / Serial / Power / Voltage / Ratio ให้ตรงกับเอกสารขาย'),
        ('2', 'ตรวจสภาพภายนอก สี รอยกระแทก หน้าแปลน ขาแท่น เพลา Keyway และอุปกรณ์ประกอบ'),
        ('3', 'ตรวจการประกอบ Motor + Gear: Bolt, Coupling, Adapter, Alignment และความแน่นของจุดยึด'),
        ('4', 'ตรวจอัตราทดเกียร์ ทิศทางการหมุน และความผิดปกติของ Output Shaft'),
        ('5', 'Run Test แบบ No-load: เสียงผิดปกติ การสั่น อุณหภูมิ และการรั่วซึมของน้ำมัน'),
        ('6', 'วัดกระแสมอเตอร์และเทียบ Nameplate / ตรวจความสมดุล 3 เฟสเมื่อเกี่ยวข้อง'),
        ('7', 'ตรวจ Terminal Box, Wiring, Grounding, Cable, Plug และ Controller/Drive เมื่อเกี่ยวข้อง'),
        ('8', 'ตรวจชนิดน้ำมันเกียร์ ปริมาณน้ำมัน ระดับน้ำมัน และยืนยันว่าเติมแล้วสำหรับรุ่นที่ต้องเติม'),
        ('9', 'ตรวจรูปถ่ายหลักฐาน: Nameplate, กระแส, เสียง/Run Test และภาพประกอบก่อนแพ็กสินค้า'),
        ('10', 'ตรวจ Packing, Label, QR/Barcode, Warranty และเอกสารแนบก่อนส่งมอบ'),
    ]

def _motor_qc_role_config(role: str):
    role = str(role or '').strip().lower()
    configs = {
        'warehouse': {
            'role': 'warehouse',
            'label': 'Warehouse / Packing',
            'title': 'Warehouse เตรียมสินค้าเรียบร้อย',
            'popup_message': 'โปรดตรวจสอบความถูกต้องอีกครั้ง หากเรียบร้อยแล้ว โปรดลงชื่อและกด Approve',
            'status': 'warehouse_prepared',
            'button': 'Approve เตรียมสินค้าเรียบร้อย',
            'to': ['wiroj@synergy-as.com', 'sas04@synergy-as.com' , 'psungpan@gmail.com'],
            'cc': ['tanai@synergy-as.com','Chottanin@synergy-as.com' , 'sas06@synergy-as.com' , 'traiwit@synergy-as.com' , 'kongkiat@synergy-as.com'],
            'subject_tpl': 'เตรียมสินค้าเรียบร้อยแล้ว QC Report OR No. : {qr_no} : {company_name} โปรดทำการ QC สินค้าตามรายการที่กำหนด',
            'body_tpl': 'เตรียมสินค้าเรียบร้อยแล้ว QC Report OR No. : {qr_no} : {company_name}  โปรดทำการ QC สินค้าตามรายการที่กำหนด\n\nด้วยความเคารพ \nTanai B. ( Eas )',
        },
        'qc': {
            'role': 'qc',
            'label': 'QC Inspector',
            'title': 'QC Inspector ตรวจสินค้าเรียบร้อย',
            'popup_message': 'โปรดตรวจสอบความถูกต้องอีกครั้ง หากเสร็จสิ้น โปรดลงชื่อและกด Approve',
            'status': 'qc_approved',
            'button': 'Approve QC เรียบร้อย',
            'requires_qc_checklist': True,
            'to': ['tanai@synergy-as.com', 'sas04@synergy-as.com'],
            'cc': ['Chottanin@synergy-as.com' , 'sas06@synergy-as.com' ,'psungpan@gmail.com' , 'traiwit@synergy-as.com' , 'kongkiat@synergy-as.com'],
            'subject_tpl': 'QC สินค้าเรียบร้อยแล้ว OR No. : {qr_no} : {company_name} โปรดทำการจัดส่งสินค้าตามรายการที่กำหนดได้เลย',
            'body_tpl': 'QC สินค้าเรียบร้อยแล้ว OR No. : {qr_no} : {company_name}  โปรดทำการจัดส่งสินค้าตามรายการที่กำหนดได้เลย\n\nด้วยความเคารพ \n\nQC Inspector',
        },
    }
    return configs.get(role)


@app.route('/')
def home():
    return render_template('index.html')

@app.route('/Video_Training.html')
@app.route('/video-training')
def video_training():
    return render_template('Video_Training.html')

# ✅ SAS Training Level 1 + Quiz
# Flask ไม่เปิดไฟล์ใน templates ด้วย URL ตรง ๆ ต้องประกาศ route ให้ render_template
@app.route('/sas_training_level1.html')
@app.route('/sas-training-level1')
def sas_training_level1():
    return render_template('sas_training_level1.html')

@app.route('/sas_quiz.html')
@app.route('/sas-quiz')
def sas_quiz():
    return render_template('sas_quiz.html')

@app.route("/healthz")
def healthz():
    return "ok", 200

@app.route('/stock')
def stock_page():
    return render_template('stock.html')

@app.route('/api/stock')
def stock_api():
    try:
        rows_main = _load_stock_rows_cached()

        try:
            rows_supplier = _load_supplier_rows_cached()
        except Exception:
            rows_supplier = []

        rows = (rows_main or []) + (rows_supplier or [])
        return jsonify({"ok": True, "count": len(rows), "rows": rows})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ✅ Supplier upload page
@app.route("/supplier-upload", methods=["GET", "POST"])
def supplier_upload_public():
    if request.method == "GET":
        key = (request.args.get("key") or "").strip()
        if key != SUPPLIER_UPLOAD_PASS:
            return "Unauthorized", 403

        session["supplier_upload_ok"] = True
        return render_template("stock_upload_supplier.html")

    # ===== POST =====
    if not session.get("supplier_upload_ok"):
        return "Unauthorized", 403

    f = request.files.get("file")
    if not f or f.filename.strip() == "":
        return "No file uploaded", 400

    if not f.filename.lower().endswith(".xlsx"):
        return "Invalid file type (ต้องเป็น .xlsx เท่านั้น)", 400

    # ✅ อัปโหลดขึ้น R2 แทน Firebase Storage
    r2_upload_fileobj(
        f,
        SUPPLIER_BLOB_NAME,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # ล้าง cache
    with _supplier_lock:
        _supplier_cache["mtime"] = None
        _supplier_cache["rows"] = []

    session.pop("supplier_upload_ok", None)
    return redirect("/stock")


@app.route('/check-status')
def check_status_page():
    return render_template('check_status.html')


@app.route('/api/check-status')
def check_status_api():
    try:
        rows = _load_check_rows_cached()
        return jsonify({"ok": True, "count": len(rows), "rows": rows})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route('/admin-motor-qc', methods=['GET'])
def admin_motor_qc():
    return render_template(
        'admin_motor_qc.html',
        product_types=PRODUCT_TYPES,
        default_qr_no=_next_motor_qc_no()
    )


@app.route('/admin-motor-qc/generate', methods=['POST'])
def admin_motor_qc_generate():
    try:
        qr_no = (request.form.get('qr_no') or _next_motor_qc_no()).strip().upper()
        company_name = (request.form.get('company_name') or '').strip()
        booking_pdf_file = request.files.get('booking_pdf')

        product_types = request.form.getlist('item_product_type[]')
        models = request.form.getlist('item_model[]')
        quantities = request.form.getlist('item_qty[]')
        assemblies = request.form.getlist('item_assembly[]')

        items = []
        row_count = min(max(len(models), len(product_types), len(quantities), len(assemblies)), 30)
        for i in range(row_count):
            model = str(models[i] if i < len(models) else '').strip()
            product_type = str(product_types[i] if i < len(product_types) else '').strip()
            if not model and not product_type:
                continue
            qty_raw = quantities[i] if i < len(quantities) else '1'
            try:
                qty = max(1, int(float(str(qty_raw).strip() or '1')))
            except Exception:
                qty = 1
            assembly = assemblies[i] if i < len(assemblies) else 'ไม่ประกอบ'
            assembly = 'ประกอบ' if str(assembly).strip() == 'ประกอบ' else 'ไม่ประกอบ'
            items.append({
                'no': len(items) + 1,
                'product_type': product_type,
                'model': model,
                'qty': qty,
                'assembly': assembly,
            })

        if (not qr_no) or (not company_name) or (not items) or any(not x.get('product_type') or not x.get('model') for x in items):
            return render_template(
                'admin_motor_qc.html',
                product_types=PRODUCT_TYPES,
                default_qr_no=qr_no or _next_motor_qc_no(),
                error='กรุณากรอก QR No., บริษัท, ประเภทสินค้า, Model และจำนวน ให้ครบอย่างน้อย 1 รายการ',
                old=request.form
            ), 400

        if not booking_pdf_file or not booking_pdf_file.filename:
            return render_template(
                'admin_motor_qc.html',
                product_types=PRODUCT_TYPES,
                default_qr_no=qr_no or _next_motor_qc_no(),
                error='กรุณาอัปโหลดไฟล์ใบจอง PDF เพื่อแนบคู่กับ QC-GEARMOTOR PRE-CHECK DOCUMENT',
                old=request.form
            ), 400

        original_booking_filename = booking_pdf_file.filename.strip()
        if not original_booking_filename.lower().endswith('.pdf'):
            return render_template(
                'admin_motor_qc.html',
                product_types=PRODUCT_TYPES,
                default_qr_no=qr_no or _next_motor_qc_no(),
                error='ไฟล์ใบจองต้องเป็น PDF เท่านั้น',
                old=request.form
            ), 400

        booking_pdf_bytes = booking_pdf_file.read()
        if not booking_pdf_bytes:
            return render_template(
                'admin_motor_qc.html',
                product_types=PRODUCT_TYPES,
                default_qr_no=qr_no or _next_motor_qc_no(),
                error='ไฟล์ใบจอง PDF ว่างหรืออ่านไฟล์ไม่ได้',
                old=request.form
            ), 400

        try:
            booking_page_count = _pdf_page_count(booking_pdf_bytes)
            if booking_page_count <= 0:
                raise RuntimeError('ไม่พบหน้าใน PDF')
        except Exception as e:
            return render_template(
                'admin_motor_qc.html',
                product_types=PRODUCT_TYPES,
                default_qr_no=qr_no or _next_motor_qc_no(),
                error=f'ไฟล์ใบจอง PDF ไม่ถูกต้องหรือรวมไฟล์ไม่ได้: {e}',
                old=request.form
            ), 400

        job_key = _safe_firebase_key(qr_no)
        safe_booking_filename = secure_filename(original_booking_filename) or f'{job_key}_booking.pdf'
        booking_pdf_key = f"motor_qc_jobs/{job_key}/booking_{safe_booking_filename}"
        booking_pdf_url = r2_upload_bytes(booking_pdf_bytes, booking_pdf_key, 'application/pdf')

        now_th = datetime.datetime.utcnow() + datetime.timedelta(hours=7)
        form_url = url_for('form', motor_qc_job=job_key, _external=True)
        approval_urls = _motor_qc_approval_urls(job_key)
        product_type_summary = _product_type_summary(items)
        now_str = now_th.strftime('%Y-%m-%d %H:%M:%S')

        job = {
            'job_key': job_key,
            'qr_no': qr_no,
            'company_name': company_name,
            'product_type': product_type_summary,
            'product_types': sorted(list({x.get('product_type') for x in items if x.get('product_type')})),
            'items': items,
            'form_url': form_url,
            'approval_urls': approval_urls,
            'item_count': len(items),
            'booking_pdf_filename': original_booking_filename,
            'booking_pdf_key': booking_pdf_key,
            'booking_pdf_url': booking_pdf_url,
            'booking_pdf_page_count': booking_page_count,
            'created_by': 'Admin Motor',
            'created_at': now_str,
            'status': 'admin_generated',
            'approvals': {
                'admin': {
                    'name': 'Admin SAS04',
                    'approved_at': now_str,
                    'role_label': 'Admin Motor SAS04',
                    'method': 'Approve Generate PDF',
                }
            }
        }

        # ✅ บันทึกข้อมูลไว้ก่อน เพื่อให้ QR / Barcode Scan แล้วเปิดข้อมูลได้ทันที
        motor_qc_jobs_ref.child(job_key).set(job)

        # ✅ สร้าง PDF พร้อม Admin Approve + QR สำหรับ Warehouse และ QC Inspector
        pdf_bytes = _build_motor_qc_precheck_pdf_bytes(job)

        # ✅ อัปโหลดขึ้น R2 เพื่อให้มีประวัติและเปิดย้อนหลังได้
        pdf_key = f"motor_qc_jobs/{job_key}.pdf"
        pdf_url = r2_upload_bytes(pdf_bytes, pdf_key, 'application/pdf')
        motor_qc_jobs_ref.child(job_key).update({'pdf_url': pdf_url})

        # ✅ ส่งอีเมลถึง Warehouse พร้อมแนบ PDF ฉบับเดียวกับที่ดาวน์โหลด
        subject = f"เอกสาร QC Report OR No. : {qr_no} : {company_name} โปรดเตรียมสินค้าตามรายการที่กำหนด"
        body = f"เอกสาร QC Report OR No. : {qr_no} : {company_name} เรียบร้อยแล้ว โปรดเตรียมสินค้าตามรายการที่กำหนด\n\nด้วยความเคารพ \nAdmin SAS04"
        mail_result = _send_workflow_email(
            to_list=['tanai@synergy-as.com'],
            cc_list=['Chottanin@synergy-as.com' , 'psungpan@gmail.com','wiroj@synergy-as.com' , 'sas04@synergy-as.com' ,'sas06@synergy-as.com' , 'traiwit@synergy-as.com' , 'kongkiat@synergy-as.com'],
            subject=subject,
            body=body,
            attachment_bytes=pdf_bytes,
            attachment_name=f"{qr_no}_QC_Motor_Precheck.pdf",
            stage='admin_to_warehouse',
            job_key=job_key,
        )
        if not mail_result.get('ok'):
            # ไม่หยุดการดาวน์โหลด PDF แต่บันทึก/แสดงใน Render Logs + Firebase last_mail_status แล้ว
            print(f"[WORKFLOW WARNING] Admin generated PDF but email failed: {mail_result.get('error')}", flush=True)

        return send_file(
            io.BytesIO(pdf_bytes),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"{qr_no}_QC_Motor_Precheck.pdf"
        )
    except Exception as e:
        return f"เกิดข้อผิดพลาดในการสร้างเอกสาร QC-Motor: {e}", 500


@app.route('/motor-qc-approve/<role>/<job_key>', methods=['GET', 'POST'])
def motor_qc_department_approve(role, job_key):
    cfg = _motor_qc_role_config(role)
    if not cfg:
        return "ไม่พบประเภทการ Approve", 404

    safe_key = _safe_firebase_key(job_key)
    job = motor_qc_jobs_ref.child(safe_key).get() or None
    if not job:
        return "ไม่พบเอกสาร QC-Motor", 404

    if request.method == 'POST':
        approver_name = (request.form.get('approver_name') or '').strip()
        signature_data = (request.form.get('signature_data') or '').strip()

        if not approver_name or not signature_data.startswith('data:image'):
            return render_template(
                'motor_qc_approve.html',
                job=job,
                cfg=cfg,
                role=cfg['role'],
                approved=False,
                error='กรุณากรอกชื่อและวาดลายเซ็นก่อนกด Approve',
                qc_checkpoints=_motor_qc_checkpoints()
            ), 400

        qc_checklist_payload = {}
        if cfg.get('requires_qc_checklist'):
            required_checks = _motor_qc_checkpoints()
            missing = []
            for no, _text in required_checks:
                status = (request.form.get(f'qc_status_{no}') or '').strip().upper()
                note = (request.form.get(f'qc_note_{no}') or '').strip()
                if status not in ('OK', 'NG'):
                    missing.append(no)
                else:
                    qc_checklist_payload[str(no)] = {'status': status, 'note': note}
            if missing:
                return render_template(
                    'motor_qc_approve.html',
                    job=job,
                    cfg=cfg,
                    role=cfg['role'],
                    approved=False,
                    error='กรุณาเลือก OK หรือ NG ให้ครบทุกข้อก่อนกด Approve เพื่อให้ PDF ล่าสุดมีเครื่องหมาย X ลงช่องที่ถูกต้อง',
                    qc_checkpoints=required_checks
                ), 400

        now_th = datetime.datetime.utcnow() + datetime.timedelta(hours=7)
        now_str = now_th.strftime('%Y-%m-%d %H:%M:%S')

        approvals = job.get('approvals') or {}
        approvals[cfg['role']] = {
            'name': approver_name,
            'signature_data': signature_data,
            'approved_at': now_str,
            'role_label': cfg['label'],
            'method': 'QR Approval',
        }
        if cfg['role'] == 'warehouse':
            wh_nos = request.form.getlist('warehouse_item_no[]')
            wh_qtys = request.form.getlist('warehouse_item_prepared_qty[]')
            wh_notes = request.form.getlist('warehouse_item_note[]')
            wh_items = []
            items_list = job.get('items') or []
            item_by_no = {str((it or {}).get('no')): it for it in items_list if isinstance(it, dict)}
            for idx, no in enumerate(wh_nos[:30]):
                no_s = str(no or idx + 1).strip() or str(idx + 1)
                qty_raw = wh_qtys[idx] if idx < len(wh_qtys) else ''
                note = (wh_notes[idx] if idx < len(wh_notes) else '').strip()
                try:
                    prepared_qty = max(0, int(float(str(qty_raw).strip() or '0')))
                except Exception:
                    prepared_qty = 0
                wh_items.append({'no': no_s, 'prepared_qty': prepared_qty, 'note': note})
                if no_s in item_by_no:
                    item_by_no[no_s]['warehouse_prepared_qty'] = prepared_qty
                    item_by_no[no_s]['warehouse_note'] = note
            approvals[cfg['role']]['items'] = wh_items
            job['items'] = items_list
        if qc_checklist_payload:
            approvals[cfg['role']]['qc_checklist'] = qc_checklist_payload
        job['approvals'] = approvals
        job['status'] = cfg['status']
        job['updated_at'] = now_str

        update_payload = {
            'approvals': approvals,
            'status': cfg['status'],
            'updated_at': now_str,
        }
        if cfg['role'] == 'warehouse':
            update_payload['items'] = job.get('items') or []
        motor_qc_jobs_ref.child(safe_key).update(update_payload)

        pdf_bytes = _build_motor_qc_precheck_pdf_bytes(job)
        pdf_key = f"motor_qc_jobs/{safe_key}.pdf"
        pdf_url = r2_upload_bytes(pdf_bytes, pdf_key, 'application/pdf')
        motor_qc_jobs_ref.child(safe_key).update({'pdf_url': pdf_url})

        qr_no = job.get('qr_no') or safe_key
        company_name = job.get('company_name') or '-'
        subject = cfg['subject_tpl'].format(qr_no=qr_no, company_name=company_name)
        body = cfg['body_tpl'].format(qr_no=qr_no, company_name=company_name)

        mail_result = _send_workflow_email(
            to_list=cfg['to'],
            cc_list=cfg['cc'],
            subject=subject,
            body=body,
            attachment_bytes=pdf_bytes,
            attachment_name=f"{qr_no}_QC_Motor_Precheck.pdf",
            stage=f"{cfg['role']}_approval",
            job_key=safe_key,
        )

        return render_template(
            'motor_qc_approve.html',
            job=job,
            cfg=cfg,
            role=cfg['role'],
            approved=True,
            pdf_url=pdf_url,
            mail_result=mail_result,
            mail_error=None if mail_result.get('ok') else mail_result.get('error'),
            qc_checkpoints=_motor_qc_checkpoints(),
        )

    return render_template(
        'motor_qc_approve.html',
        job=job,
        cfg=cfg,
        role=cfg['role'],
        approved=False,
        qc_checkpoints=_motor_qc_checkpoints(),
    )


@app.route('/login', methods=['GET', 'POST'])
def login():
    next_url = _safe_next_url(request.values.get('next') or session.get('after_login_url', ''))

    if request.method == 'POST':
        ip = _get_client_ip()

        # Rate limit check
        if not check_login_rate_limit(ip):
            _queue_alert(
                f"LOGIN RATE LIMIT EXCEEDED — IP BLOCKED\n"
                f"IP: {ip}\nUA: {request.headers.get('User-Agent','')[:200]}"
            )
            BLOCKED_IPS.add(ip)
            return render_template('login.html', error=True, blocked=True, next_url=next_url)

        employee_id = (request.form.get('employee_id') or '').strip().upper()
        allowed_ids = list(INSPECTOR_MAP.keys())

        if employee_id not in allowed_ids:
            record_failed_login(ip, employee_id)
            return render_template('login.html', error=True, next_url=next_url)

        session['employee_id'] = employee_id
        session['inspector_name'] = INSPECTOR_MAP.get(employee_id, employee_id)
        session['just_logged_in'] = True

        target = _safe_next_url(request.form.get('next') or session.pop('after_login_url', ''))
        if target:
            return redirect(target)
        return redirect(url_for('form'))

    return render_template('login.html', next_url=next_url)


@app.route('/form')
def form():
    if 'employee_id' not in session:
        # ✅ ถ้า QC สแกน QR จากเอกสารก่อน login ให้จำ URL เดิมไว้ แล้วกลับมาหน้านี้หลัง login
        if request.query_string:
            session['after_login_url'] = request.full_path.rstrip('?')
        return redirect(url_for('login', next=request.full_path.rstrip('?') if request.query_string else ''))

    just_logged_in = session.pop('just_logged_in', False)

    prefill_job = None
    motor_qc_job_key = (request.args.get('motor_qc_job') or '').strip()
    if motor_qc_job_key:
        prefill_job = motor_qc_jobs_ref.child(_safe_firebase_key(motor_qc_job_key)).get() or None

    return render_template(
        'form.html',
        employee_id=session['employee_id'],
        welcome=just_logged_in,
        prefill_job=prefill_job,
        prefill_job_json=json.dumps(prefill_job, ensure_ascii=False) if prefill_job else 'null'
    )


@app.route('/submit', methods=['POST'])
def submit():
    try:
        if not (request.content_type or '').startswith('multipart/form-data'):
            return "Invalid Content-Type", 400

        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        serial = f"SAS{timestamp}"

        def upload_image(file, field_name):
            """อัปโหลดรูปขึ้น R2 คืน public URL"""
            if file and file.filename:
                filename = secure_filename(file.filename)
                key = f"qc_images/{serial}_{field_name}_{filename}"
                file_bytes = file.stream.read()
                return r2_upload_bytes(file_bytes, key, file.content_type or "image/jpeg")
            return None

        inspector = session.get('inspector_name') or request.form.get('inspector')

        # ✅ ข้อมูลจากเอกสาร QC-Motor ที่ Admin Motor Generate ไว้
        motor_qc_job_key = request.form.get('motor_qc_job_key')
        motor_qc_qr_no = request.form.get('motor_qc_qr_no')
        assembly_status = request.form.get('assembly_status')

        # ============================================================
        # ✅ New mode: QC หลายรายการใน 1 Order
        # ============================================================
        qc_item_nos = request.form.getlist('qc_item_no[]')
        qc_items = []

        if qc_item_nos:
            product_types = request.form.getlist('qc_item_product_type[]')
            models = request.form.getlist('qc_item_model[]')
            assemblies = request.form.getlist('qc_item_assembly[]')
            quantities = request.form.getlist('qc_item_qty[]')
            gear_ratios = request.form.getlist('qc_item_gear_ratio[]')
            motor_currents = request.form.getlist('qc_item_motor_current[]')
            gear_sounds = request.form.getlist('qc_item_gear_sound[]')
            warranties = request.form.getlist('qc_item_warranty[]')
            oil_types = request.form.getlist('qc_item_oil_type[]')
            oil_liters_list = request.form.getlist('qc_item_oil_liters[]')
            servo_motor_models = request.form.getlist('qc_item_servo_motor_model[]')
            servo_drive_models = request.form.getlist('qc_item_servo_drive_model[]')

            file_keys = {
                'motor_current_img': 'ภาพค่ากระแส',
                'gear_sound_img': 'ภาพเสียงเกียร์',
                'assembly_img': 'ภาพประกอบหน้างาน',
                'rfks_nameplate_motor_img': 'Name plate : Motor',
                'rfks_nameplate_gear_img': 'Name plate : Gear',
                'controller_img': 'ภาพ Controller',
                'servo_motor_img': 'ภาพ Servo Motor',
                'servo_drive_img': 'ภาพ Servo Drive',
                'cable_wire_img': 'ภาพ Cable Wire',
            }

            for idx, raw_no in enumerate(qc_item_nos[:30], start=1):
                no = str(raw_no or idx).strip() or str(idx)
                product_type_i = product_types[idx-1] if idx-1 < len(product_types) else ''
                model_i = models[idx-1] if idx-1 < len(models) else ''
                assembly_i = assemblies[idx-1] if idx-1 < len(assemblies) else ''
                qty_raw = quantities[idx-1] if idx-1 < len(quantities) else '1'
                try:
                    qty_i = max(1, int(float(str(qty_raw).strip() or '1')))
                except Exception:
                    qty_i = 1
                item_images = {}
                for file_key in file_keys.keys():
                    # ชื่อ field ฝั่ง form เช่น qc_item_motor_current_img_1
                    field_name = f"qc_item_{file_key}_{idx}"
                    item_images[file_key] = upload_image(
                        request.files.get(field_name),
                        f"item{no}_{file_key}"
                    )

                item = {
                    'no': no,
                    'product_type': product_type_i,
                    'model': model_i,
                    'qty': qty_i,
                    'assembly': assembly_i,
                    'gear_ratio': gear_ratios[idx-1] if idx-1 < len(gear_ratios) else '',
                    'motor_current': motor_currents[idx-1] if idx-1 < len(motor_currents) else '',
                    'gear_sound': gear_sounds[idx-1] if idx-1 < len(gear_sounds) else '',
                    'warranty': warranties[idx-1] if idx-1 < len(warranties) else '',
                    'oil_type': oil_types[idx-1] if idx-1 < len(oil_types) else '',
                    'oil_liters': oil_liters_list[idx-1] if idx-1 < len(oil_liters_list) else '',
                    'oil_filled': 'เติมแล้ว' if request.form.get(f'qc_item_oil_filled_{idx}') else 'ยังไม่เติม',
                    'acdc_parts': request.form.getlist(f'qc_item_acdc_parts_{idx}[]'),
                    'servo_motor_model': servo_motor_models[idx-1] if idx-1 < len(servo_motor_models) else '',
                    'servo_drive_model': servo_drive_models[idx-1] if idx-1 < len(servo_drive_models) else '',
                    'images': item_images,
                }
                qc_items.append(item)

            product_type = _product_type_summary(qc_items)
            motor_nameplate = "\n".join([
                f"Item {it.get('no')}: {it.get('model','')} x {it.get('qty', 1)} ({it.get('product_type','')} / {it.get('assembly','')})"
                for it in qc_items
            ])
            motor_current = "\n".join([f"Item {it.get('no')}: {it.get('motor_current','')} A" for it in qc_items if it.get('motor_current')])
            gear_ratio = "\n".join([f"Item {it.get('no')}: {it.get('gear_ratio','')}" for it in qc_items if it.get('gear_ratio')])
            gear_sound = "\n".join([f"Item {it.get('no')}: {it.get('gear_sound','')} dB" for it in qc_items if it.get('gear_sound')])
            warranty = "\n".join([f"Item {it.get('no')}: {it.get('warranty','')} เดือน" for it in qc_items if it.get('warranty')])
            oil_type = ''
            oil_liters = ''
            oil_filled = ''
            acdc_parts = []
            servo_motor_model = ''
            servo_drive_model = ''
            images = {}
            motor_qc_item_no = ''

        else:
            # ============================================================
            # Legacy mode: ฟอร์มเดิม 1 รายการ
            # ============================================================
            product_type = request.form.get('product_type')
            motor_nameplate = request.form.get('motor_nameplate')
            motor_current = request.form.get('motor_current')
            gear_ratio = request.form.get('gear_ratio')
            gear_sound = request.form.get('gear_sound')
            warranty = request.form.get('warranty')
            oil_type = request.form.get('oil_type')
            oil_liters = request.form.get('oil_liters')
            oil_filled = 'เติมแล้ว' if request.form.get('oil_filled') else 'ยังไม่เติม'
            acdc_parts = request.form.getlist('acdc_parts')
            servo_motor_model = request.form.get('servo_motor_model')
            servo_drive_model = request.form.get('servo_drive_model')
            motor_qc_item_no = request.form.get('motor_qc_item_no')

            images = {
                'motor_current_img': upload_image(request.files.get('motor_current_img'), 'motor_current'),
                'gear_sound_img': upload_image(request.files.get('gear_sound_img'), 'gear_sound'),
                'assembly_img': upload_image(request.files.get('assembly_img'), 'assembly'),
                'controller_img': upload_image(request.files.get('controller_img'), 'controller'),
                'servo_motor_img': upload_image(request.files.get('servo_motor_img'), 'servo_motor'),
                'servo_drive_img': upload_image(request.files.get('servo_drive_img'), 'servo_drive'),
                'cable_wire_img': upload_image(request.files.get('cable_wire_img'), 'cable_wire'),
                'rfks_nameplate_motor_img': upload_image(request.files.get('rfks_nameplate_motor_img'), 'rfks_nameplate_motor'),
                'rfks_nameplate_gear_img': upload_image(request.files.get('rfks_nameplate_gear_img'), 'rfks_nameplate_gear'),
            }

        or_no = request.form.get('or_no')
        company_name = request.form.get('company_name')

        payload = {
            'serial': serial,
            'or_no': or_no,
            'company_name': company_name,
            'product_type': product_type,
            'motor_nameplate': motor_nameplate,
            'motor_current': motor_current,
            'gear_ratio': gear_ratio,
            'gear_sound': gear_sound,
            'warranty': warranty,
            'inspector': inspector,
            'oil_type': oil_type,
            'oil_liters': oil_liters,
            'oil_filled': oil_filled,
            'acdc_parts': acdc_parts,
            'servo_motor_model': servo_motor_model,
            'servo_drive_model': servo_drive_model,
            'motor_qc_job_key': motor_qc_job_key,
            'motor_qc_qr_no': motor_qc_qr_no,
            'motor_qc_item_no': motor_qc_item_no,
            'assembly_status': assembly_status,
            'images': images,
            'qc_items': qc_items,
            'date': datetime.datetime.now().strftime('%Y-%m-%d')
        }

        ref.child(serial).set(payload)

        # ✅ อัปเดตสถานะรายการใน QC-Motor Job ว่าถูก Scan/Submit แล้ว
        if motor_qc_job_key:
            try:
                job_key_safe = _safe_firebase_key(motor_qc_job_key)
                now_th = (datetime.datetime.utcnow() + datetime.timedelta(hours=7)).strftime('%Y-%m-%d %H:%M:%S')
                motor_qc_jobs_ref.child(job_key_safe).child('last_qc_serial').set(serial)
                motor_qc_jobs_ref.child(job_key_safe).child('last_qc_submit_at').set(now_th)
                if qc_items:
                    status_obj = {str(it.get('no')): {
                        'qc_serial': serial,
                        'submitted_at': now_th,
                        'product_type': it.get('product_type'),
                        'model': it.get('model'),
                        'qty': it.get('qty'),
                    } for it in qc_items}
                    motor_qc_jobs_ref.child(job_key_safe).child('qc_items_status').update(status_obj)
            except Exception as _e:
                print(f"⚠️ update motor_qc_jobs failed: {_e}", flush=True)

        def background_finalize():
            try:
                data = ref.child(serial).get()
                image_urls, image_labels = _collect_qc_report_images(data)

                # ✅ สร้าง PDF รองรับ qc_items หลายรายการ
                pdf_stream = create_qc_pdf(data, image_urls=image_urls, image_labels=image_labels)
                pdf_stream.seek(0)

                # ✅ อัปโหลด PDF ขึ้น R2
                pdf_key = f"qc_reports/{serial}.pdf"
                pdf_url = r2_upload_bytes(pdf_stream.read(), pdf_key, "application/pdf")

                # ✅ สร้าง QR จาก public URL จริง
                qr_stream = generate_qr_code(serial, pdf_url)
                qr_key = f"qr_codes/{serial}.png"
                qr_url = r2_upload_bytes(qr_stream.read(), qr_key, "image/png")

                ref.child(serial).update({
                    'qc_pdf_url': pdf_url,
                    'qr_png_url': qr_url
                })
                print(f"✅ PDF + QR สำหรับ {serial} สร้างเสร็จ", flush=True)
            except Exception as e:
                print(f"❌ Error in background finalize: {e}", flush=True)

        threading.Thread(target=background_finalize).start()
        return redirect(url_for('success', serial=serial))

    except Exception as e:
        return f"เกิดข้อผิดพลาด: {e}", 400


@app.route('/success')
def success():
    serial = request.args.get('serial', '')
    data = ref.child(serial).get()
    return render_template('success.html',
                           serial_number=serial,
                           qc_url=data.get("qc_pdf_url", "#"),
                           qr_url=data.get("qr_png_url", "#"))


@app.route('/download/<serial_number>')
def download_pdf(serial_number):
    report_data = ref.child(serial_number).get()
    if not report_data:
        return "ไม่พบรายงาน", 404

    image_urls, image_labels = _collect_qc_report_images(report_data)
    pdf_stream = create_qc_pdf(report_data, image_urls=image_urls, image_labels=image_labels)
    pdf_stream.seek(0)
    return send_file(pdf_stream,
                     as_attachment=True,
                     download_name=f"{serial_number}_QC_Report.pdf",
                     mimetype='application/pdf')


@app.route('/qr/<serial_number>')
def generate_qr(serial_number):
    # ✅ ถ้ามีลิงก์ PDF แล้ว ให้ QR ชี้ไปที่ PDF จริง
    report_data = ref.child(serial_number).get() or {}
    pdf_url = report_data.get("qc_pdf_url")

    if pdf_url:
        link = pdf_url
    else:
        # fallback ถ้ายังไม่เสร็จ
        link = f"https://sas-qc-gearmotor.onrender.com/autodownload/{serial_number}"

    qr = qrcode.QRCode(version=1, box_size=10, border=4)
    qr.add_data(link)
    qr.make(fit=True)
    img = qr.make_image(fill_color='black', back_color='white')
    qr_stream = io.BytesIO()
    img.save(qr_stream, format='PNG')
    qr_stream.seek(0)
    return send_file(qr_stream, mimetype='image/png', download_name=f"{serial_number}_QR.png")


@app.route('/autodownload/<serial_number>')
def autodownload(serial_number):
    report_data = ref.child(serial_number).get()
    if not report_data:
        return "ไม่พบรายงาน", 404

    image_urls, image_labels = _collect_qc_report_images(report_data)
    pdf_stream = create_qc_pdf(report_data, image_urls=image_urls, image_labels=image_labels)
    pdf_stream.seek(0)
    return send_file(pdf_stream,
                     as_attachment=True,
                     download_name=f"{serial_number}_QC_Report.pdf",
                     mimetype='application/pdf')



# ============================================================
# 🎓 SAS Training Level 2 / Level 3 Routes
# ============================================================
# หมายเหตุ: ต้องวางไฟล์ sas_training_level2.html, sas_quiz2.html,
# sas_training_level3.html, sas_quiz3.html ไว้ในโฟลเดอร์ templates

@app.route('/sas_training_level2.html')
@app.route('/sas-training-level2')
def sas_training_level2():
    return render_template('sas_training_level2.html')


@app.route('/sas_quiz2.html')
@app.route('/sas-quiz2')
def sas_quiz2():
    return render_template('sas_quiz2.html')


@app.route('/sas_training_level3.html')
@app.route('/sas-training-level3')
def sas_training_level3():
    return render_template('sas_training_level3.html')


@app.route('/sas_quiz3.html')
@app.route('/sas-quiz3')
def sas_quiz3():
    return render_template('sas_quiz3.html')


@app.route('/sas_training_4series.html')
@app.route('/sas-training-4series')
def sas_training_4series():
    return render_template('sas_training_4series.html')


# ============================================================
# 🏆 SAS Training Quiz Database API
# ใช้ Firebase Realtime Database เป็นฐานกลาง เพื่อเช็คข้ามเครื่อง
# ============================================================
training_quiz_ref = db.reference('/training_quiz')


def _digits_only(v):
    return ''.join(ch for ch in str(v or '') if ch.isdigit())


def _norm_text(v):
    return str(v or '').strip().lower().replace(' ', '')


def _training_level_key(level):
    s = str(level or '').strip().lower()
    if s in ('1', 'level1', 'quiz1', 'l1'):
        return 'level1'
    if s in ('2', 'level2', 'quiz2', 'l2'):
        return 'level2'
    if s in ('3', 'level3', 'quiz3', 'l3'):
        return 'level3'
    return s or 'level1'


def _training_total_score_for(level_key):
    return 40 if level_key == 'level3' else 30


def _training_pass_score_for(level_key):
    if level_key == 'level3':
        return 34
    return 25


def _training_doc_id(name, phone):
    n = _norm_text(name)[:80]
    p = _digits_only(phone)[-10:]
    return f'n_{n}_p_{p}' if (n or p) else 'unknown'


def _training_row_score(row):
    try:
        return int(row.get('score') or row.get('correct') or 0)
    except Exception:
        return 0


def _training_row_passed(level_key, row):
    return bool(row.get('passed')) or _training_row_score(row) >= _training_pass_score_for(level_key)


def _find_training_record(level_key, name='', phone=''):
    name_n = _norm_text(name)
    phone_n = _digits_only(phone)
    phone_last = phone_n[-10:] if phone_n else ''
    rows_obj = training_quiz_ref.child(level_key).get() or {}
    if not isinstance(rows_obj, dict):
        return None

    # 1) match เบอร์โทรก่อน เพราะข้ามเครื่องแม่นสุด
    if phone_last:
        for row in rows_obj.values():
            if not isinstance(row, dict):
                continue
            rp = _digits_only(row.get('phone'))
            if rp and rp[-10:] == phone_last:
                return row

    # 2) match ชื่อแบบตัดช่องว่าง
    if name_n:
        for row in rows_obj.values():
            if not isinstance(row, dict):
                continue
            rn = _norm_text(row.get('name'))
            if rn and rn == name_n:
                return row

    return None


@app.route('/api/training-quiz/save', methods=['POST'])
def training_quiz_save():
    try:
        data = request.get_json(silent=True) or {}
        level_key = _training_level_key(data.get('level') or data.get('quiz_level') or data.get('level_key'))
        name = str(data.get('name') or data.get('fullName') or data.get('fullname') or '').strip()
        pos = str(data.get('pos') or data.get('position') or data.get('team') or '').strip()
        phone = str(data.get('phone') or data.get('tel') or data.get('mobile') or '').strip()
        score = int(data.get('score') or data.get('correct') or 0)
        total = int(data.get('total') or _training_total_score_for(level_key))
        passed = bool(data.get('passed')) or score >= _training_pass_score_for(level_key)
        now = (datetime.datetime.utcnow() + datetime.timedelta(hours=7)).strftime('%Y-%m-%d %H:%M:%S')

        doc_id = _training_doc_id(name, phone)
        old = training_quiz_ref.child(level_key).child(doc_id).get() or {}
        old_score = int(old.get('score') or 0) if isinstance(old, dict) else 0

        payload = {
            'level': level_key,
            'name': name,
            'pos': pos,
            'phone': phone,
            'score': max(score, old_score),
            'total': total,
            'passed': passed or bool(old.get('passed')) if isinstance(old, dict) else passed,
            'updated_at': now,
            'created_at': old.get('created_at') if isinstance(old, dict) and old.get('created_at') else now,
        }
        training_quiz_ref.child(level_key).child(doc_id).set(payload)
        return jsonify({'ok': True, 'record': payload})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/training-quiz/check-level1', methods=['POST'])
def training_quiz_check_level1():
    try:
        data = request.get_json(silent=True) or {}
        row = _find_training_record('level1', data.get('name', ''), data.get('phone', ''))
        passed = bool(row) and _training_row_passed('level1', row)
        return jsonify({'ok': True, 'passed': passed, 'record': row or None})
    except Exception as e:
        return jsonify({'ok': False, 'passed': False, 'error': str(e)}), 500


@app.route('/api/training-quiz/check-level2', methods=['POST'])
def training_quiz_check_level2():
    try:
        data = request.get_json(silent=True) or {}
        row = _find_training_record('level2', data.get('name', ''), data.get('phone', ''))
        passed = bool(row) and _training_row_passed('level2', row)
        return jsonify({'ok': True, 'passed': passed, 'record': row or None})
    except Exception as e:
        return jsonify({'ok': False, 'passed': False, 'error': str(e)}), 500


@app.route('/api/training-quiz/check-level3-prereq', methods=['POST'])
def training_quiz_check_level3_prereq():
    try:
        data = request.get_json(silent=True) or {}
        name = data.get('name', '')
        phone = data.get('phone', '')
        row1 = _find_training_record('level1', name, phone)
        row2 = _find_training_record('level2', name, phone)
        passed1 = bool(row1) and _training_row_passed('level1', row1)
        passed2 = bool(row2) and _training_row_passed('level2', row2)
        return jsonify({
            'ok': True,
            'passed': bool(passed1 and passed2),
            'passed_level1': bool(passed1),
            'passed_level2': bool(passed2),
            'level1': row1 or None,
            'level2': row2 or None,
        })
    except Exception as e:
        return jsonify({'ok': False, 'passed': False, 'error': str(e)}), 500


def _training_overall_rows():
    bucket = {}

    def ident(row):
        return _training_doc_id(row.get('name', ''), row.get('phone', ''))

    for level_key in ('level1', 'level2', 'level3'):
        rows_obj = training_quiz_ref.child(level_key).get() or {}
        if not isinstance(rows_obj, dict):
            continue
        for row in rows_obj.values():
            if not isinstance(row, dict):
                continue
            key = ident(row)
            if not key or key == 'unknown':
                continue
            if key not in bucket:
                bucket[key] = {
                    'name': row.get('name', ''),
                    'pos': row.get('pos', ''),
                    'phone': row.get('phone', ''),
                    'level1': 0,
                    'level2': 0,
                    'level3': 0,
                    'passed_level1': False,
                    'passed_level2': False,
                    'passed_level3': False,
                    'date': '',
                }
            item = bucket[key]
            item['name'] = item.get('name') or row.get('name', '')
            item['pos'] = item.get('pos') or row.get('pos', '')
            item['phone'] = item.get('phone') or row.get('phone', '')
            score = _training_row_score(row)
            item[level_key] = max(int(item.get(level_key) or 0), score)
            item['passed_' + level_key] = bool(item.get('passed_' + level_key)) or _training_row_passed(level_key, row)
            d = row.get('updated_at') or row.get('created_at') or ''
            if d > item.get('date', ''):
                item['date'] = d

    rows = []
    for item in bucket.values():
        item['total_score'] = int(item.get('level1') or 0) + int(item.get('level2') or 0) + int(item.get('level3') or 0)
        item['total'] = item['total_score']
        item['total_possible'] = 100
        rows.append(item)
    rows.sort(key=lambda x: (x.get('total_score', 0), x.get('level3', 0), x.get('date', '')), reverse=True)
    return rows


@app.route('/api/training-quiz/leaderboard/total')
def training_quiz_leaderboard_total():
    try:
        return jsonify({'ok': True, 'level': 'total', 'rows': _training_overall_rows()[:200]})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e), 'rows': []}), 500


@app.route('/api/training-quiz/leaderboard/<level>')
def training_quiz_leaderboard(level):
    try:
        if str(level or '').strip().lower() in ('total', 'overall', 'all'):
            return jsonify({'ok': True, 'level': 'total', 'rows': _training_overall_rows()[:200]})
        level_key = _training_level_key(level)
        rows_obj = training_quiz_ref.child(level_key).get() or {}
        rows = []
        if isinstance(rows_obj, dict):
            for row in rows_obj.values():
                if not isinstance(row, dict):
                    continue
                rows.append({
                    'name': row.get('name', ''),
                    'pos': row.get('pos', ''),
                    'phone': row.get('phone', ''),
                    'score': int(row.get('score') or 0),
                    'total': int(row.get('total') or _training_total_score_for(level_key)),
                    'passed': _training_row_passed(level_key, row),
                    'date': row.get('updated_at') or row.get('created_at') or '',
                })
        rows.sort(key=lambda x: (x.get('score', 0), x.get('date', '')), reverse=True)
        return jsonify({'ok': True, 'level': level_key, 'rows': rows[:200]})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e), 'rows': []}), 500



# ============================================================
# 📊 SALES DASHBOARD — Routes + API
# ============================================================
sales_ref = db.reference("/sales_quotations")

@app.route('/sales-dashboard')
@app.route('/sales_dashboard.html')
def sales_dashboard():
    return render_template('sales_dashboard.html')

@app.route('/api/sales/records', methods=['GET'])
def api_sales_records():
    try:
        data = sales_ref.get() or {}
        records = list(data.values())
        records.sort(key=lambda x: x.get('date', ''), reverse=True)
        return jsonify({'ok': True, 'records': records})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

# -- PDF Parser helpers --
def _fix_num(s):
    if not s: return 0.0
    s = str(s).strip()
    import re as _re
    s = _re.sub(r'(\d)\s+(\d)', r'\1\2', s)
    s = _re.sub(r'(\d)\s+,',    r'\1,',   s)
    s = _re.sub(r',\s+(\d)',    r',\1',   s)
    s = s.replace(',', '').strip()
    try: return float(s)
    except: return 0.0

def _detect_format(tables, lines):
    """ตรวจว่า PDF เป็น format ใด: TW (QMO6905) หรือ SAS (RS/AP/NM)"""
    for tbl in tables:
        if not tbl: continue
        h = ' '.join(str(c or '') for c in tbl[0]).upper()
        if 'SALE PERSON' in h: return 'TW'
        if 'SALES' in h:       return 'SAS'
    for line in lines:
        if 'Sale Person' in line: return 'TW'
    return 'SAS'

def _parse_tw_format(tables, lines, text):
    """Parse QMO6905 / TW format (Sale Person header, Grand Total, multiline items)"""
    r = {'id':'','customer':'','date':'','saleShort':'XX',
         'exVat':0.0,'vat':0.0,'total':0.0,'items':[],
         'delivery':'-','payment':'-','warranty':'-','saleEngineer':''}

    m = re.search(r'Quotation\s*No\.?\s*:?\s*([\w\-]+)', text, re.I)
    if m: r['id'] = m.group(1).strip()

    m = re.search(r'Date:\s*(\d{1,2})[\/](\d{1,2})[\/](\d{2,4})', text)
    if m:
        d, mo, y = m.group(1), m.group(2).zfill(2), m.group(3)
        yr = int(y)
        if len(y) == 2: yr = (2500+yr)-543
        elif yr > 2500: yr -= 543
        r['date'] = f"{yr}-{mo}-{d.zfill(2)}"

    m = re.search(r'To:\s*(.+)', text)
    if m:
        cust = re.sub(r'\(cid:\d+\)', '', m.group(1)).strip()
        if len(cust) > 2: r['customer'] = cust

    for tbl in tables:
        if not tbl: continue
        h_up = ' '.join(str(c or '') for c in tbl[0]).upper()
        if 'SALE PERSON' not in h_up: continue
        if len(tbl) > 1:
            row1 = tbl[1]
            v0 = str(row1[0] or '').strip()
            if v0: r['saleEngineer'] = v0
            for cell in row1:
                cv = str(cell or '').strip()
                if re.search(r'\d+\s*days?', cv, re.I): r['delivery'] = cv; break
        break

    for line in lines:
        m = re.search(r'Grand\s+Total\s+([\d\s,\.]+)', line)
        if m:
            v = _fix_num(m.group(1))
            if v > 100: r['total'] = v
    for line in lines:
        m = re.search(r'[Vv]at\s*7%\s+([\d\s,\.]+)', line)
        if m:
            v = _fix_num(m.group(1))
            if v > 0: r['vat'] = v; break
    for line in lines:
        m = re.match(r'^Total\s+([\d\s,\.]+)$', line)
        if m:
            v = _fix_num(m.group(1))
            if v > 100: r['exVat'] = v
    if r['exVat'] == 0 and r['total'] > 0 and r['vat'] > 0:
        r['exVat'] = round(r['total'] - r['vat'], 2)
    elif r['exVat'] == 0 and r['total'] > 0 and r['vat'] == 0:
        r['exVat'] = r['total']  # Vat 0% → exVat = total

    for tbl in tables:
        if not tbl: continue
        h_up = ' '.join(str(c or '') for c in tbl[0]).upper()
        if 'SALE PERSON' not in h_up: continue
        for row in tbl[3:]:
            if not row: continue
            item_cell = str(row[0] or '').strip()
            item_nos  = [x.strip() for x in item_cell.split('\n') if x.strip().isdigit()]
            if not item_nos: continue
            codes = [x.strip() for x in str(row[3] or '').split('\n')
                     if x.strip() and not x.startswith('**') and '(cid:' not in x]
            qtys  = [x.strip() for x in str(row[6] or '').split('\n') if x.strip().isdigit()]
            units = [x.strip() for x in str(row[8] or '').split('\n')
                     if x.strip() and re.search(r'[\d,]', x)]
            tots  = [x.strip() for x in str(row[9] or '').split('\n')
                     if x.strip() and re.search(r'[\d,]', x)]
            all_desc = [x.strip() for x in str(row[4] or '').split('\n')
                        if x.strip() and not x.startswith('**') and '(cid:' not in x]
            n = len(codes)
            desc_groups = [[] for _ in range(n)]
            cur = 0
            for dl in all_desc:
                if cur < n-1 and desc_groups[cur] and not re.match(r'^[a-z]', dl):
                    cur += 1
                desc_groups[cur].append(dl)
            for idx in range(n):
                code = codes[idx] if idx < len(codes) else ''
                desc = ' '.join(desc_groups[idx]) if idx < len(desc_groups) else ''
                qty  = int(qtys[idx]) if idx < len(qtys) else 1
                unit = _fix_num(units[idx]) if idx < len(units) else 0.0
                itot = _fix_num(tots[idx])  if idx < len(tots)  else 0.0
                if unit > 0 or itot > 0:
                    r['items'].append({'code':code,'desc':desc[:120],'qty':qty,'unit':unit,'total':itot})
        break
    return r

def _parse_sas_pdf(file_storage):
    """Universal parser: AP/NM (QTY/Total Price/VAT 7%) + RS (Q'TY/Total include vat/Vat 7%)"""
    import io as _io
    pdf_bytes = file_storage.read()
    file_storage.seek(0)
    with pdfplumber.open(_io.BytesIO(pdf_bytes)) as pdf:
        # อ่านทุกหน้า — totals มักอยู่หน้าสุดท้าย
        all_text, all_tables = [], []
        for pg in pdf.pages:
            t = pg.extract_text(x_tolerance=3, y_tolerance=3) or ''
            all_text.append(t)
            all_tables.extend(pg.extract_tables() or [])
        text   = '\n'.join(all_text)
        lines  = [l.strip() for l in text.split('\n') if l.strip()]
        tables = all_tables
    # ── ตรวจ format แล้ว route ไปยัง parser ที่เหมาะสม ──
    fmt = _detect_format(tables, lines)
    if fmt == 'TW':
        return _parse_tw_format(tables, lines, text)

    r = {'id':'','customer':'','date':'','saleShort':'XX',
         'exVat':0.0,'vat':0.0,'total':0.0,
         'items':[],'delivery':'-','payment':'-','warranty':'-'}
    # Quotation ID
    m = re.search(r'Quotation\s*NO\.?\s*:?\s*(QM[\w\-]+)', text, re.I)
    if m: r['id'] = m.group(1).strip()
    # Sale/Warranty/Delivery/Payment -- header table only
    BLACKLIST = {'ITEM','CODE','DESCRIPTION','QTY',"Q'TY",'UNITPRICE','TOTAL',
                 'SALES','WARRANTY','SHIPPING','METHOD','JOB','DELIVERY','DATE','PAYMENT','TERMS'}
    for tbl in tables:
        if not tbl or len(tbl) < 2: continue
        h = [str(c or '').strip().upper().replace('\n',' ') for c in tbl[0]]
        if 'SALES' not in h: continue
        for row in tbl[1:]:
            if not row: continue
            v0 = str(row[0] or '').strip()
            if v0 and v0 not in BLACKLIST:
                if re.match(r'^[A-Z]{2,4}$', v0):
                    r['saleShort'] = v0
                else:
                    r['saleEngineer'] = v0  # full name (PPY format)
                if len(row) > 1 and row[1]: r['warranty'] = str(row[1]).strip()
                for cell in row:
                    cv = str(cell or '').strip()
                    if re.search(r'Within|days|weeks|วัน', cv, re.I): r['delivery']=cv; break
                for cell in reversed(row):
                    cv = str(cell or '').strip()
                    if cv and cv not in BLACKLIST and cv != '-' and len(cv) < 40:
                        r['payment']=cv; break
                break
        if r['saleShort'] != 'XX': break
    # Date — รองรับทุก format
    _MONTHS = {'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,
               'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12}
    def _parse_yr(y):
        yr=int(y)
        if len(y)==2: return (2500+yr)-543
        return yr-543 if yr>2500 else yr
    _date_found = False
    # DD/MM/YYYY or D/M/YY
    _m = re.search(r'Date\.?\s*:?\s*(\d{1,2})[\/](\d{1,2})[\/](\d{2,4})', text)
    if _m:
        r['date'] = f"{_parse_yr(_m.group(3))}-{_m.group(2).zfill(2)}-{_m.group(1).zfill(2)}"
        _date_found = True
    # "May 7, 2026" or "7 May 2026"
    if not _date_found:
        _m = re.search(r'Date\.?\s*:?\s*([A-Za-z]{3,9})\s+(\d{1,2}),?\s*(\d{4})', text)
        if _m and _m.group(1).lower()[:3] in _MONTHS:
            r['date'] = f"{_m.group(3)}-{str(_MONTHS[_m.group(1).lower()[:3]]).zfill(2)}-{_m.group(2).zfill(2)}"
            _date_found = True
    if not _date_found:
        _m = re.search(r'(\d{1,2})\s+([A-Za-z]{3,9})\s+(\d{4})', text)
        if _m and _m.group(2).lower()[:3] in _MONTHS:
            r['date'] = f"{_m.group(3)}-{str(_MONTHS[_m.group(2).lower()[:3]]).zfill(2)}-{_m.group(1).zfill(2)}"
    # Customer -- อ่านจาก PDF (แต่ชื่อไทยมักเพี้ยนเพราะ font encoding)
    # ใช้เป็น fallback เท่านั้น, ชื่อจริงดึงจากชื่อไฟล์ใน upload handler
    for i, line in enumerate(lines):
        if re.match(r'^To\s*:', line):
            inline = re.sub(r'^To\s*:\s*', '', line).strip()
            if inline and len(inline) > 3 and not re.match(r'^(Email|Tel|Fax|Khun|คุณ)', inline, re.I):
                r['customer'] = inline; break
            for j in range(i+1, min(i+6, len(lines))):
                nl = lines[j].strip()
                if nl and not re.match(r'^(Email|Tel|Fax|AP|NM|TW|RS|CA|LINE|SALES|Khun|คุณ|\d)', nl, re.I):
                    r['customer'] = nl; break
            break

    # Sale Engineer -- ดึงจากบรรทัดใกล้ท้าย (หลัง "Sale Engineer" header)
    r['saleEngineer'] = ''
    for i, line in enumerate(lines):
        if re.search(r'Sale\s+Engineer|Sales\s+Engineer|Sales\s+Supervisor', line, re.I):
            # ชื่ออยู่บรรทัดถัดไป หรือบรรทัดเดียวกันกับ "Ms./Mr. Approve By"
            for j in range(i+1, min(i+3, len(lines))):
                nl = lines[j].strip()
                if nl and re.match(r'^(Ms\.|Mr\.|Mrs\.|[A-Z][a-z])', nl):
                    # ตัด Manager/Approve By ออก
                    name = re.sub(r'\s*Mr\.Traiwit\s+Luengthong.*$', '', nl, flags=re.I).strip()
                    name = re.sub(r'\s*(Sale|General)\s+Manager.*$', '', name, flags=re.I).strip()
                    if name and len(name) > 3:
                        r['saleEngineer'] = name
                        break
            break
    # Totals -- both "Total Price" and "Total include vat"
    grand_re = re.compile(r'Total\s+(?:include\s+vat|Price|with\s+vat)\s+([\d\s,\.]+)', re.I)
    tc = []
    for line in lines:
        m = grand_re.search(line)
        if m:
            v = _fix_num(m.group(1))
            if v > 100: tc.append(v)
    if tc: r['total'] = tc[-1]
    # VAT -- both "VAT 7%" and "Vat 7%"
    for line in lines:
        m = re.search(r'[Vv][Aa][Tt]\s*7%\s+([\d\s,\.]+)', line)
        if m:
            v = _fix_num(m.group(1))
            if v > 0: r['vat'] = v; break
    # exVat
    for line in lines:
        m = re.match(r'^\d+\s+Total(?:\s+Price)?\s+([\d\s,\.]+)$', line)
        if m:
            v = _fix_num(m.group(1))
            if v > 0: r['exVat'] = v; break
    if r['exVat'] == 0 and r['total'] > 0 and r['vat'] > 0:
        r['exVat'] = round(r['total'] - r['vat'], 2)
    elif r['exVat'] == 0 and r['total'] > 0 and r['vat'] == 0:
        r['exVat'] = r['total']  # Vat 0% → exVat = total
    # Items -- both QTY and Q'TY
    for tbl in tables:
        if not tbl: continue
        h = [str(c or '').strip().upper() for c in tbl[0]]
        if not ('ITEM' in h and ('QTY' in h or "Q'TY" in h)): continue
        ci = h.index('ITEM')
        cc = h.index('CODE')        if 'CODE'        in h else -1
        cd = h.index('DESCRIPTION') if 'DESCRIPTION' in h else -1
        cq = (h.index("Q'TY") if "Q'TY" in h else
              h.index('QTY')   if 'QTY'  in h else
              h.index('QTY.')  if 'QTY.' in h else -1)
        # "UNIT PRICE" หรือ "UNITPRICE" → ใช้ตัวแรกที่เจอ
        cu = next((i for i,v in enumerate(h) if 'UNIT' in v and 'PRICE' in v), -1)
        ct = h.index('TOTAL')       if 'TOTAL'       in h else -1
        for row in tbl[1:]:
            if not row or len(row) <= ci: continue
            item_no = str(row[ci] or '').strip()
            if not item_no or not item_no.isdigit(): continue
            code  = str(row[cc] or '').strip() if cc >= 0 else ''
            desc  = str(row[cd] or '').strip() if cd >= 0 else ''
            qty_s = str(row[cq] or '').strip() if cq >= 0 else '1'
            unit  = _fix_num(str(row[cu] or '')) if cu >= 0 else 0.0
            itot  = _fix_num(str(row[ct] or '')) if ct >= 0 else 0.0
            if unit > 0 or itot > 0:
                qty = int(qty_s) if qty_s.isdigit() else 1
                full_desc = f"{code} -- {desc}".strip(' -') if code else desc
                r['items'].append({'code':code or f'ITEM{item_no}',
                                   'desc':full_desc[:120],'qty':qty,
                                   'unit':unit,'total':itot})
        break
    return r


@app.route('/api/sales/upload', methods=['POST'])
def api_sales_upload():
    try:
        file = request.files.get('file')
        if not file or not file.filename.endswith('.pdf'):
            return jsonify({'ok': False, 'error': 'ต้องเป็นไฟล์ PDF'}), 400

        filename = secure_filename(file.filename)

        # ── Parse PDF บน server ── (แม่นยำ 99%)
        parsed = _parse_sas_pdf(file)

        # ── Fallback อัจฉริยะจากชื่อไฟล์ ──
        def _date_from_parts(d, mo, y_str):
            yr = int(y_str)
            if len(y_str) == 2: yr = (2500 + yr) - 543  # 69 → 2026
            elif yr > 2500:     yr -= 543               # พ.ศ. 4 หลัก
            return f"{yr}-{mo.zfill(2)}-{d.zfill(2)}"

        KNOWN_SALES_LIST = ['TWS','TW','NM','AP','CA','BW','TL','NR','CK','SI','RS','AW','PK','BS','PL','PPY','KSR']

        def _parse_fname(fname):
            """Universal filename parser: returns (qid, sale_short, customer)"""
            base = re.sub(r'\.pdf$','',fname,flags=re.I)
            base = re.sub(r'__\d{1,2}-\d{2}-\d{4}_*$','',base)
            base = re.sub(r'\s*\(\d{1,2}-\d{2}-\d{2,4}\)\s*$','',base).strip()
            parts = base.split('-')
            sale_idx = -1
            sale_sh  = ''
            for i, p in enumerate(parts):
                if p.strip().upper() in KNOWN_SALES_LIST:
                    sale_idx = i; sale_sh = p.strip().upper(); break
            if sale_idx == -1:
                return base, 'XX', ''
            pre   = parts[:sale_idx]
            after = parts[sale_idx+1:]
            # QMO6905 style: digit ก่อน sale → QID = pre เท่านั้น
            old_style = sale_idx > 1 and re.match(r'^\d+$', parts[sale_idx-1])
            if old_style:
                qid = '-'.join(pre)
                cust_parts = after
            else:
                serials, cust_start = [], 0
                for i, p in enumerate(after):
                    if re.match(r'^\d+$',p.strip()) or re.match(r'^R\d+$',p.strip(),re.I):
                        serials.append(p.strip()); cust_start = i+1
                    else: break
                qid = '-'.join(pre+[sale_sh]+serials)
                cust_parts = after[cust_start:]
            cust = '-'.join(cust_parts).replace('_',' ')
            cust = re.sub(r'\s+',' ',cust).strip(' -')
            return qid, sale_sh, cust

        def _customer_from_filename(fname):
            _, _, cust = _parse_fname(fname)
            return cust

        if not parsed['id']:
            qm = re.match(r'^(QM[\w]+-[\w]+-\d+(?:-R\d+)?)', filename, re.I)
            parsed['id'] = qm.group(1) if qm else re.sub(r'\.pdf$','',filename,flags=re.I)

        # ดึง QID, saleShort, customer จากชื่อไฟล์ (แม่นกว่า PDF เสมอ)
        fn_qid, fn_sale, fn_cust = _parse_fname(filename)
        if fn_sale != 'XX':   parsed['saleShort'] = fn_sale
        if fn_cust:            parsed['customer']  = fn_cust
        if not parsed['id'] and fn_qid: parsed['id'] = fn_qid
        if not parsed['date']:
            dm = re.search(r'\((\d{1,2})-(\d{2})-(\d{2,4})\)', filename)
            if dm:
                parsed['date'] = _date_from_parts(dm.group(1),dm.group(2),dm.group(3))
            else:
                dm2 = re.search(r'(\d{2})-(\d{2})-(\d{4})', filename)
                if dm2:
                   parsed['date'] = _date_from_parts(dm2.group(1),dm2.group(2),dm2.group(3))
                else:
                   parsed['date'] = datetime.datetime.utcnow().strftime('%Y-%m-%d')

        # ชื่อลูกค้า: ชื่อไฟล์ชัวร์กว่า PDF (Thai font encoding เพี้ยน)
        fn_cust = _customer_from_filename(filename)
        if fn_cust:
            parsed['customer'] = fn_cust  # ใช้จากชื่อไฟล์เสมอ


        # ── ป้องกันซ้ำ: เช็คทั้ง Quotation ID และ filename ──
        safe_id = parsed['id'].replace('/','_').replace('.','_')
        existing = sales_ref.child(safe_id).get()
        if existing:
            # มี Quotation ID นี้แล้ว — return ข้อมูลเดิม พร้อม flag duplicate
            return jsonify({
                'ok': True,
                'duplicate': True,
                'record': existing,
                'message': f"ใบเสนอราคา {parsed['id']} มีในระบบแล้ว (อัพโหลดเมื่อ {existing.get('uploaded_at','?')})"
            })

        # ── อัพโหลด PDF ขึ้น R2 ──
        file.seek(0)
        pdf_bytes = file.read()
        r2_key  = f"sales_pdf/{filename}"
        pdf_url = r2_upload_bytes(pdf_bytes, r2_key, 'application/pdf')

        # ── บันทึกลง Firebase ──
        record = {
            'id':         parsed['id'],
            'filename':   filename,
            'pdf_url':    pdf_url,
            'saleShort':  parsed['saleShort'],
            'customer':   parsed['customer'],
            'date':       parsed['date'],
            'month':      parsed['date'][:7] if parsed['date'] else '',
            'exVat':      float(parsed['exVat']),
            'vat':        float(parsed['vat']),
            'total':      float(parsed['total']),
            'items':      parsed['items'],
            'delivery':   parsed['delivery'],
            'payment':    parsed['payment'],
            'warranty':   parsed['warranty'],
            'saleEngineer': parsed.get('saleEngineer', ''),
            'status':     None,
            'note':       '',
            'uploaded_at': datetime.datetime.utcnow().isoformat(),
        }
        sales_ref.child(safe_id).set(record)
        return jsonify({'ok': True, 'duplicate': False, 'record': record})
    except Exception as e:
        import traceback
        return jsonify({'ok': False, 'error': str(e), 'trace': traceback.format_exc()}), 500

@app.route('/api/sales/record/<record_id>', methods=['PATCH'])
def api_sales_patch(record_id):
    try:
        body    = request.get_json(force=True) or {}
        safe_id = record_id.replace('/', '_').replace('.', '_')
        ref     = sales_ref.child(safe_id)

        update_fields = {}

        # status update
        if 'status' in body:
            update_fields['status'] = body['status']
            update_fields['status_updated_at'] = body.get('status_updated_at')

        # note: append ลง notes[] array (history log)
        if 'add_note' in body:
            note_text = str(body['add_note']).strip()
            if note_text:
                existing = ref.get() or {}
                notes = existing.get('notes') or []
                if not isinstance(notes, list): notes = []
                notes.append({
                    'rev':  len(notes) + 1,
                    'text': note_text,
                    'ts':   datetime.datetime.utcnow().isoformat() + 'Z'
                })
                update_fields['notes'] = notes

        if not update_fields:
            return jsonify({'ok': False, 'error': 'nothing to update'}), 400

        ref.update(update_fields)
        return jsonify({'ok': True, 'notes': update_fields.get('notes')})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/sales/record/<record_id>', methods=['DELETE'])
def api_sales_delete(record_id):
    try:
        safe_id = record_id.replace('/', '_').replace('.', '_')
        sales_ref.child(safe_id).delete()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/sales/clear-all', methods=['POST'])
def api_sales_clear_all():
    """ล้างข้อมูลทั้งหมดใน Firebase (ใช้เมื่อต้องการ reset)"""
    try:
        sales_ref.delete()
        return jsonify({'ok': True, 'message': 'Cleared all sales records'})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True)
